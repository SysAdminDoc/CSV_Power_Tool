#!/usr/bin/env python3
"""
CSV Power Tool
Professional-grade CSV file combiner and processor.
Merge, filter, transform, deduplicate, and export CSV data with full control.
"""

import sys
import csv
import re
import json
import threading
import time
import locale
from pathlib import Path
from datetime import datetime, timezone
from dataclasses import dataclass, field
from typing import Callable
from tkinter import filedialog, StringVar, BooleanVar, IntVar, END

APP_NAME = "CSV Power Tool"
APP_VERSION = "3.1.0"
SUPPORTED_INPUT_SUFFIXES = {".csv", ".tsv", ".txt", ".xlsx", ".parquet"}
SUPPORTED_TEXT_SUFFIXES = {".csv", ".tsv", ".txt"}
SUPPORTED_OUTPUT_SUFFIXES = {".csv", ".tsv", ".txt", ".xlsx", ".parquet"}

try:
    locale.setlocale(locale.LC_COLLATE, "")
except locale.Error:
    pass


GUI_IMPORT_ERROR = None
try:
    import customtkinter as ctk
except ImportError as exc:
    GUI_IMPORT_ERROR = exc

    class _MissingCTkFrame:
        def __init__(self, *args, **kwargs):
            raise RuntimeError("customtkinter is required for GUI mode")

    class _MissingCTk:
        CTkFrame = _MissingCTkFrame

    ctk = _MissingCTk()

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except ImportError:
    DND_AVAILABLE = False
else:
    DND_AVAILABLE = GUI_IMPORT_ERROR is None


# ══════════════════════════════════════════════════════════════════════════════
# THEME CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

COLORS = {
    "bg_dark": "#020617",
    "bg_secondary": "#0f172a",
    "bg_tertiary": "#1e293b",
    "bg_input": "#0f172a",
    "bg_hover": "#334155",
    "border": "#334155",
    "border_light": "#475569",
    "text_primary": "#f8fafc",
    "text_secondary": "#94a3b8",
    "text_muted": "#64748b",
    "accent_green": "#22c55e",
    "accent_green_hover": "#16a34a",
    "accent_blue": "#60a5fa",
    "accent_blue_hover": "#3b82f6",
    "accent_purple": "#a78bfa",
    "accent_purple_hover": "#8b5cf6",
    "accent_orange": "#f97316",
    "accent_red": "#ef4444",
    "accent_cyan": "#22d3ee",
    "success": "#22c55e",
    "warning": "#f59e0b",
    "error": "#ef4444",
    "tab_active": "#1e293b",
    "tab_inactive": "#0f172a",
}

FONTS = {
    "title": ("Segoe UI", 24, "bold"),
    "heading": ("Segoe UI", 14, "bold"),
    "subheading": ("Segoe UI", 12, "bold"),
    "body": ("Segoe UI", 12),
    "small": ("Segoe UI", 11),
    "mono": ("Consolas", 11),
}


# ══════════════════════════════════════════════════════════════════════════════
# DATA MODELS
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class ProcessingConfig:
    """Configuration for CSV processing."""
    # Column selection
    columns_mode: str = "all"  # "all", "select", "exclude"
    selected_columns: list = field(default_factory=list)
    column_mapping: dict = field(default_factory=dict)  # {original: renamed}
    column_order: list = field(default_factory=list)

    # Sorting
    sort_enabled: bool = False
    sort_columns: list = field(default_factory=list)  # [(column, ascending), ...]
    sort_case_sensitive: bool = False
    sort_numeric_aware: bool = True

    # Deduplication
    dedupe_enabled: bool = True
    dedupe_columns: list = field(default_factory=list)  # Empty = all columns
    dedupe_keep: str = "first"  # "first", "last", "none"
    dedupe_fuzzy_enabled: bool = False
    dedupe_fuzzy_threshold: int = 90
    dedupe_aggregate_mode: str = "none"  # "none", "max", "min", "sum", "concat"
    dedupe_aggregate_separator: str = "; "

    # Filtering
    filters: list = field(default_factory=list)  # [(column, operator, value), ...]
    filter_logic: str = "and"  # "and", "or"

    # Transformations
    trim_whitespace: bool = True
    case_transform: str = "none"  # "none", "upper", "lower", "title"
    empty_value: str = ""  # Replace empty cells with this

    # Header normalization
    header_normalize: str = "none"  # "none", "trim", "lowercase", "snake_case"
    strip_bom: bool = True

    # Advanced transforms (per-column)
    column_transforms: list = field(default_factory=list)  # [(column, transform_type, *args)]
    # transform_type: "trim", "upper", "lower", "title", "replace", "regex_replace",
    #                  "split", "merge", "compute"

    # Schema unification
    schema_mode: str = "union"  # "union", "intersection", "first_file"

    # Output
    output_delimiter: str = ","
    output_encoding: str = "utf-8"
    output_quoting: str = "minimal"  # "minimal", "all", "nonnumeric", "none"
    include_header: bool = True
    line_ending: str = "auto"  # "auto", "unix", "windows"
    streaming_enabled: bool = True


@dataclass 
class ProcessingStats:
    """Statistics from processing."""
    files_processed: int = 0
    files_skipped: int = 0
    total_rows_read: int = 0
    rows_filtered: int = 0
    duplicates_removed: int = 0
    final_row_count: int = 0
    unique_columns: int = 0
    errors: list = field(default_factory=list)


# ══════════════════════════════════════════════════════════════════════════════
# CSV PROCESSING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class CSVEngine:
    """Core CSV processing engine."""

    FILTER_OPERATORS = {
        "equals": lambda v, t: str(v).lower() == str(t).lower(),
        "not_equals": lambda v, t: str(v).lower() != str(t).lower(),
        "contains": lambda v, t: str(t).lower() in str(v).lower(),
        "not_contains": lambda v, t: str(t).lower() not in str(v).lower(),
        "starts_with": lambda v, t: str(v).lower().startswith(str(t).lower()),
        "ends_with": lambda v, t: str(v).lower().endswith(str(t).lower()),
        "is_empty": lambda v, t: not str(v).strip(),
        "is_not_empty": lambda v, t: bool(str(v).strip()),
        "greater_than": lambda v, t: CSVEngine._numeric_compare(v, t, ">"),
        "less_than": lambda v, t: CSVEngine._numeric_compare(v, t, "<"),
        "greater_than_or_equal": lambda v, t: CSVEngine._numeric_compare(v, t, ">="),
        "less_than_or_equal": lambda v, t: CSVEngine._numeric_compare(v, t, "<="),
        "between": lambda v, t: CSVEngine._between(v, t),
        "fuzzy": lambda v, t: CSVEngine._fuzzy_match(v, t),
        "in_list": lambda v, t: str(v).strip().lower() in [x.strip().lower() for x in str(t).split(",")],
        "not_in_list": lambda v, t: str(v).strip().lower() not in [x.strip().lower() for x in str(t).split(",")],
        "regex": lambda v, t: bool(re.search(t, str(v), re.IGNORECASE)),
    }

    @staticmethod
    def _parse_number(value):
        try:
            return float(str(value).replace(",", "").strip())
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_datetime(value):
        """Parse common date/time strings and normalize aware values to UTC."""
        text = str(value).strip()
        if not text:
            return None

        normalized = text[:-1] + "+00:00" if text.endswith("Z") else text
        try:
            parsed = datetime.fromisoformat(normalized)
        except ValueError:
            parsed = None

        if parsed is None:
            formats = [
                "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d",
                "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S",
                "%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M", "%d/%m/%Y %H:%M",
                "%b %d %Y", "%B %d %Y", "%d %b %Y", "%d %B %Y",
            ]
            for fmt in formats:
                try:
                    parsed = datetime.strptime(text.replace(",", ""), fmt)
                    break
                except ValueError:
                    continue

        if parsed is None:
            try:
                from email.utils import parsedate_to_datetime
                parsed = parsedate_to_datetime(text)
            except (TypeError, ValueError, IndexError):
                return None

        if parsed.tzinfo is not None:
            parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
        return parsed

    @staticmethod
    def _numeric_compare(v, t, op):
        v_num = CSVEngine._parse_number(v)
        t_num = CSVEngine._parse_number(t)
        if v_num is None or t_num is None:
            return False
        if op == ">":
            return v_num > t_num
        elif op == "<":
            return v_num < t_num
        elif op == ">=":
            return v_num >= t_num
        elif op == "<=":
            return v_num <= t_num
        return False

    @staticmethod
    def _between(v, t):
        """Check if a number or date is between two bounds."""
        parts = None
        target = str(t)
        for separator in ("..", "|", ","):
            split = target.split(separator)
            if len(split) == 2:
                parts = [split[0].strip(), split[1].strip()]
                break
        if parts is None:
            return False

        val_num = CSVEngine._parse_number(v)
        lo_num = CSVEngine._parse_number(parts[0])
        hi_num = CSVEngine._parse_number(parts[1])
        if val_num is not None and lo_num is not None and hi_num is not None:
            return lo_num <= val_num <= hi_num

        val_date = CSVEngine._parse_datetime(v)
        lo_date = CSVEngine._parse_datetime(parts[0])
        hi_date = CSVEngine._parse_datetime(parts[1])
        if val_date and lo_date and hi_date:
            return lo_date <= val_date <= hi_date

        return False

    @staticmethod
    def _fuzzy_match(v, t):
        """Return True when v is similar to target. Target may be 'text|90'."""
        target = str(t).strip()
        threshold = 85
        for separator in ("|", ","):
            candidate, sep, maybe_threshold = target.rpartition(separator)
            if sep and maybe_threshold.strip().isdigit():
                target = candidate.strip()
                threshold = int(maybe_threshold.strip())
                break

        if not target:
            return False

        value = str(v).strip()
        value_lower = value.lower()
        target_lower = target.lower()
        if target_lower in value_lower or value_lower in target_lower:
            return True
        try:
            from rapidfuzz import fuzz
            score = fuzz.partial_ratio(value_lower, target_lower)
        except ImportError:
            from difflib import SequenceMatcher
            score = SequenceMatcher(None, value_lower, target_lower).ratio() * 100
        return score >= threshold

    @staticmethod
    def _normalize_header(name: str, mode: str) -> str:
        """Normalize a header name according to the mode."""
        if mode == "none":
            return name.strip()
        name = name.strip()
        # Strip BOM if present
        if name.startswith("﻿"):
            name = name[1:]
        if mode == "trim":
            return name
        if mode == "lowercase":
            return name.lower()
        if mode == "snake_case":
            # Convert to snake_case: "First Name" -> "first_name", "firstName" -> "first_name"
            import re as _re
            s = _re.sub(r'[\s\-\.]+', '_', name)  # spaces/hyphens/dots to underscore
            s = _re.sub(r'([A-Z]+)([A-Z][a-z])', r'\1_\2', s)
            s = _re.sub(r'([a-z\d])([A-Z])', r'\1_\2', s)
            s = _re.sub(r'_+', '_', s)
            return s.lower().strip('_')
        return name

    @staticmethod
    def _detect_file_params(file_path: Path) -> tuple:
        """Auto-detect encoding, delimiter, and quotechar for a file.
        Returns (encoding, delimiter, quotechar)."""
        # Try to detect encoding
        encoding = 'utf-8'
        try:
            import chardet
            with open(file_path, 'rb') as f:
                raw = f.read(min(32768, file_path.stat().st_size))
            detection = chardet.detect(raw)
            if detection and detection.get('encoding') and detection.get('confidence', 0) > 0.5:
                encoding = detection['encoding']
                # Normalize common encoding names
                enc_lower = encoding.lower().replace('-', '').replace('_', '')
                if enc_lower in ('ascii', 'utf8'):
                    encoding = 'utf-8'
                elif enc_lower in ('iso88591', 'latin1'):
                    encoding = 'latin-1'
                elif enc_lower in ('cp1252', 'windows1252'):
                    encoding = 'cp1252'
                elif enc_lower.startswith('utf16'):
                    encoding = 'utf-16'
        except ImportError:
            pass  # chardet not available, fall back to trial-and-error

        # Detect delimiter and quotechar using csv.Sniffer
        delimiter = ','
        quotechar = '"'
        for enc in [encoding, 'utf-8', 'latin-1', 'cp1252']:
            try:
                with open(file_path, 'r', encoding=enc, newline='') as f:
                    sample = f.read(8192)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
                    delimiter = dialect.delimiter
                    quotechar = dialect.quotechar or '"'
                except csv.Error:
                    # Sniffer failed, fall back to frequency counting
                    for delim in [',', '\t', ';', '|']:
                        if sample.count(delim) > sample.count(delimiter):
                            delimiter = delim
                encoding = enc
                break
            except UnicodeDecodeError:
                continue

        return encoding, delimiter, quotechar

    @staticmethod
    def _cell_to_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat()
        return str(value)

    def _discover_structured_columns(self, file_path: Path) -> list[str]:
        suffix = file_path.suffix.lower()
        norm_mode = getattr(self.config, 'header_normalize', 'none')

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("openpyxl is required for .xlsx input") from exc

            workbook = load_workbook(file_path, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                rows = sheet.iter_rows(values_only=True)
                headers = next(rows, [])
                return [
                    self._normalize_header(self._cell_to_text(col), norm_mode)
                    for col in headers
                    if self._normalize_header(self._cell_to_text(col), norm_mode)
                ]
            finally:
                workbook.close()

        if suffix == ".parquet":
            try:
                import pyarrow.parquet as pq
                schema = pq.read_schema(file_path)
                return [
                    self._normalize_header(name, norm_mode)
                    for name in schema.names
                    if self._normalize_header(name, norm_mode)
                ]
            except ImportError:
                try:
                    import polars as pl
                    frame = pl.scan_parquet(file_path)
                    return [
                        self._normalize_header(name, norm_mode)
                        for name in frame.collect_schema().names()
                        if self._normalize_header(name, norm_mode)
                    ]
                except ImportError as exc:
                    raise RuntimeError("pyarrow or polars is required for .parquet input") from exc

        return []

    def _read_structured_file(self, file_path: Path, all_columns: set, column_order: list) -> list[dict]:
        suffix = file_path.suffix.lower()
        norm_mode = getattr(self.config, 'header_normalize', 'none')
        rows = []

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(file_path, read_only=True, data_only=True)
                try:
                    sheet = workbook.active
                    row_iter = sheet.iter_rows(values_only=True)
                    raw_headers = next(row_iter, [])
                    headers = [
                        self._normalize_header(self._cell_to_text(header), norm_mode)
                        for header in raw_headers
                    ]

                    for col in headers:
                        if col and col not in all_columns:
                            all_columns.add(col)
                            column_order.append(col)

                    for values in row_iter:
                        cleaned_row = {}
                        for idx, header in enumerate(headers):
                            if not header:
                                continue
                            value = values[idx] if idx < len(values) else ""
                            cleaned_row[header] = self._cell_to_text(value).strip() if self.config.trim_whitespace else self._cell_to_text(value)
                        rows.append(cleaned_row)
                finally:
                    workbook.close()
            except Exception as exc:
                self.log(f"XX Error reading {file_path.name}: {exc}", "error")
                self.stats.files_skipped += 1
                self.stats.errors.append(f"{file_path.name}: {exc}")
                return []

        elif suffix == ".parquet":
            try:
                try:
                    import pyarrow.parquet as pq
                    table = pq.read_table(file_path)
                    records = table.to_pylist()
                except ImportError:
                    import polars as pl
                    records = pl.read_parquet(file_path).to_dicts()

                for record in records:
                    cleaned_row = {}
                    for key, value in record.items():
                        header = self._normalize_header(str(key), norm_mode)
                        if not header:
                            continue
                        if header not in all_columns:
                            all_columns.add(header)
                            column_order.append(header)
                        cleaned_row[header] = self._cell_to_text(value).strip() if self.config.trim_whitespace else self._cell_to_text(value)
                    rows.append(cleaned_row)
            except Exception as exc:
                self.log(f"XX Error reading {file_path.name}: {exc}", "error")
                self.stats.files_skipped += 1
                self.stats.errors.append(f"{file_path.name}: {exc}")
                return []

        self.stats.files_processed += 1
        self.stats.total_rows_read += len(rows)
        self.log(f"OK {file_path.name} ({len(rows):,} rows)", "success")
        return rows
    
    def __init__(self, config: ProcessingConfig, 
                 progress_callback: Callable = None,
                 log_callback: Callable = None):
        self.config = config
        self.progress_callback = progress_callback
        self.log_callback = log_callback
        self.cancelled = False
        self.stats = ProcessingStats()
    
    def log(self, message: str, level: str = "info"):
        if self.log_callback:
            self.log_callback(message, level)
    
    def update_progress(self, value: float, status: str):
        if self.progress_callback:
            self.progress_callback(value, status)
    
    def cancel(self):
        self.cancelled = True
    
    def discover_columns(self, files: list[Path]) -> list[str]:
        """Discover all unique columns across files."""
        all_columns = []
        seen = set()
        per_file_columns = []  # Track columns per file for schema modes

        norm_mode = getattr(self.config, 'header_normalize', 'none')

        for file_path in files:
            try:
                if file_path.suffix.lower() in SUPPORTED_TEXT_SUFFIXES:
                    encoding, delimiter, quotechar = self._detect_file_params(file_path)
                    with open(file_path, 'r', encoding=encoding, newline='') as f:
                        reader = csv.reader(f, delimiter=delimiter, quotechar=quotechar)
                        headers = next(reader, [])
                        file_cols = []
                        for col in headers:
                            col = self._normalize_header(col, norm_mode)
                            if col:
                                file_cols.append(col)
                                if col not in seen:
                                    all_columns.append(col)
                                    seen.add(col)
                        per_file_columns.append(set(file_cols))
                else:
                    file_cols = self._discover_structured_columns(file_path)
                    for col in file_cols:
                        if col not in seen:
                            all_columns.append(col)
                            seen.add(col)
                    per_file_columns.append(set(file_cols))
            except Exception:
                try:
                    with open(file_path, 'r', encoding='latin-1', newline='') as f:
                        reader = csv.reader(f)
                        headers = next(reader, [])
                        file_cols = []
                        for col in headers:
                            col = self._normalize_header(col, norm_mode)
                            if col and col not in seen:
                                all_columns.append(col)
                                seen.add(col)
                                file_cols.append(col)
                        per_file_columns.append(set(file_cols))
                except Exception:
                    pass

        # Apply schema unification mode
        schema_mode = getattr(self.config, 'schema_mode', 'union')
        if schema_mode == "intersection" and per_file_columns:
            common = per_file_columns[0]
            for s in per_file_columns[1:]:
                common = common & s
            all_columns = [c for c in all_columns if c in common]
        elif schema_mode == "first_file" and per_file_columns:
            first_set = per_file_columns[0]
            all_columns = [c for c in all_columns if c in first_set]

        return all_columns
    
    def process(self, input_files: list[Path], output_file: Path) -> ProcessingStats:
        """Process CSV files according to configuration."""
        self.cancelled = False
        self.stats = ProcessingStats()

        if self._can_stream(input_files, output_file):
            return self._process_streaming(input_files, output_file)

        all_rows = []
        all_columns = set()
        column_order = []
        per_file_columns = []  # Track columns per file for schema modes

        total_files = len(input_files)

        # Phase 1: Read all files
        self.log("Phase 1: Reading files...", "info")

        for idx, csv_path in enumerate(input_files):
            if self.cancelled:
                self.log("Processing cancelled", "warning")
                return self.stats

            progress = (idx / total_files) * 40
            self.update_progress(progress, f"Reading {csv_path.name}...")

            pre_cols = set(all_columns)
            rows_from_file = self._read_file(csv_path, all_columns, column_order)
            new_cols = all_columns - pre_cols
            # Track which columns this file contributed
            # We need all columns that appeared in this file's headers
            file_cols = set()
            if rows_from_file:
                file_cols = set(rows_from_file[0].keys())
            per_file_columns.append(file_cols)
            all_rows.extend(rows_from_file)

        if not all_rows:
            self.log("No data to process", "warning")
            return self.stats

        self.stats.unique_columns = len(all_columns)

        # Apply schema unification before determining final columns
        schema_mode = getattr(self.config, 'schema_mode', 'union')
        if schema_mode == "intersection" and per_file_columns:
            common = per_file_columns[0]
            for s in per_file_columns[1:]:
                common = common & s
            column_order = [c for c in column_order if c in common]
        elif schema_mode == "first_file" and per_file_columns:
            first_set = per_file_columns[0]
            column_order = [c for c in column_order if c in first_set]

        # Determine final columns
        final_columns = self._get_final_columns(column_order)
        final_columns = self._with_transform_columns(final_columns)
        
        # Phase 2: Apply filters
        if self.config.filters and not self.cancelled:
            self.update_progress(45, "Applying filters...")
            self.log(f"Phase 2: Applying {len(self.config.filters)} filter(s)...", "info")
            all_rows = self._apply_filters(all_rows)
        
        # Phase 3: Apply transformations
        if not self.cancelled:
            self.update_progress(55, "Applying transformations...")
            self.log("Phase 3: Applying transformations...", "info")
            all_rows = self._apply_transformations(all_rows, final_columns)
        
        # Phase 4: Deduplicate
        if self.config.dedupe_enabled and not self.cancelled:
            self.update_progress(65, "Removing duplicates...")
            self.log("Phase 4: Removing duplicates...", "info")
            all_rows = self._deduplicate(all_rows, final_columns)
        
        # Phase 5: Sort
        if self.config.sort_enabled and self.config.sort_columns and not self.cancelled:
            self.update_progress(80, "Sorting data...")
            self.log("Phase 5: Sorting data...", "info")
            all_rows = self._sort_rows(all_rows, final_columns)
        
        # Phase 6: Write output
        if not self.cancelled:
            self.update_progress(90, "Writing output file...")
            self.log("Phase 6: Writing output...", "info")
            self._write_output(all_rows, final_columns, output_file)
        
        self.stats.final_row_count = len(all_rows)
        self.update_progress(100, "Complete!")
        
        return self.stats

    def _can_stream(self, input_files: list[Path], output_file: Path) -> bool:
        if not getattr(self.config, "streaming_enabled", True):
            return False
        if self.config.dedupe_enabled or self.config.sort_enabled:
            return False
        if output_file.suffix.lower() not in SUPPORTED_TEXT_SUFFIXES:
            return False
        return all(path.suffix.lower() in SUPPORTED_TEXT_SUFFIXES for path in input_files)

    def _iter_text_rows(self, file_path: Path):
        if not file_path.exists():
            self.log(f"XX File not found: {file_path.name}", "error")
            self.stats.files_skipped += 1
            self.stats.errors.append(f"Not found: {file_path.name}")
            return

        norm_mode = getattr(self.config, 'header_normalize', 'none')
        detected_enc, detected_delim, detected_quote = self._detect_file_params(file_path)
        encodings_to_try = [detected_enc]
        for fallback in ['utf-8', 'latin-1', 'cp1252', 'utf-16']:
            if fallback not in encodings_to_try:
                encodings_to_try.append(fallback)

        for encoding in encodings_to_try:
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as f:
                    reader = csv.DictReader(f, delimiter=detected_delim, quotechar=detected_quote)
                    row_count = 0
                    for row in reader:
                        cleaned_row = {}
                        for key, value in row.items():
                            normalized_key = self._normalize_header(key, norm_mode) if key else ""
                            if normalized_key:
                                text_value = value or ""
                                cleaned_row[normalized_key] = text_value.strip() if self.config.trim_whitespace else text_value
                        row_count += 1
                        yield cleaned_row
                    self.stats.files_processed += 1
                    self.stats.total_rows_read += row_count
                    self.log(f"OK {file_path.name} ({row_count:,} rows, {encoding}, delim={repr(detected_delim)})", "success")
                    return
            except UnicodeDecodeError:
                continue
            except Exception as exc:
                if encoding == encodings_to_try[-1]:
                    self.log(f"XX Error reading {file_path.name}: {exc}", "error")
                    self.stats.files_skipped += 1
                    self.stats.errors.append(f"{file_path.name}: {exc}")
                continue

    def _process_streaming(self, input_files: list[Path], output_file: Path) -> ProcessingStats:
        self.log("Phase 1: Streaming rows...", "info")
        discovered_order = self.discover_columns(input_files)
        final_columns = self._with_transform_columns(self._get_final_columns(discovered_order))
        output_columns = [self.config.column_mapping.get(c, c) for c in final_columns]
        self.stats.unique_columns = len(discovered_order)

        quoting_map = {
            "minimal": csv.QUOTE_MINIMAL,
            "all": csv.QUOTE_ALL,
            "nonnumeric": csv.QUOTE_NONNUMERIC,
            "none": csv.QUOTE_NONE,
        }
        quoting = quoting_map.get(self.config.output_quoting, csv.QUOTE_MINIMAL)
        newline = "\n" if self.config.line_ending == "unix" else "\r\n" if self.config.line_ending == "windows" else ""

        output_file.parent.mkdir(parents=True, exist_ok=True)
        with open(output_file, 'w', encoding=self.config.output_encoding, newline=newline if newline else '') as f:
            writer = csv.DictWriter(
                f,
                fieldnames=output_columns,
                delimiter=self.config.output_delimiter,
                quoting=quoting,
                extrasaction='ignore'
            )
            if self.config.include_header:
                writer.writeheader()

            total_files = len(input_files)
            for idx, csv_path in enumerate(input_files):
                if self.cancelled:
                    self.log("Processing cancelled", "warning")
                    return self.stats

                progress = (idx / total_files) * 90
                self.update_progress(progress, f"Streaming {csv_path.name}...")

                for row in self._iter_text_rows(csv_path):
                    if self.config.filters and not self._row_matches_filters(row):
                        self.stats.rows_filtered += 1
                        continue

                    transformed = self._apply_transformations([row], final_columns)[0]
                    writer.writerow(transformed)
                    self.stats.final_row_count += 1

        self.update_progress(100, "Complete!")
        self.log(f"OK Saved: {output_file.name} ({self.stats.final_row_count:,} rows)", "success")
        return self.stats
    
    def _read_file(self, file_path: Path, all_columns: set, column_order: list) -> list[dict]:
        """Read a single CSV file."""
        rows = []

        if not file_path.exists():
            self.log(f"✗ File not found: {file_path.name}", "error")
            self.stats.files_skipped += 1
            self.stats.errors.append(f"Not found: {file_path.name}")
            return rows

        if file_path.suffix.lower() not in SUPPORTED_TEXT_SUFFIXES:
            return self._read_structured_file(file_path, all_columns, column_order)

        norm_mode = getattr(self.config, 'header_normalize', 'none')

        # Auto-detect encoding, delimiter, quotechar
        detected_enc, detected_delim, detected_quote = self._detect_file_params(file_path)
        encodings_to_try = [detected_enc]
        for fallback in ['utf-8', 'latin-1', 'cp1252', 'utf-16']:
            if fallback not in encodings_to_try:
                encodings_to_try.append(fallback)

        for encoding in encodings_to_try:
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as f:
                    reader = csv.DictReader(f, delimiter=detected_delim,
                                            quotechar=detected_quote)

                    if reader.fieldnames:
                        for col in reader.fieldnames:
                            col = self._normalize_header(col, norm_mode)
                            if col and col not in all_columns:
                                all_columns.add(col)
                                column_order.append(col)

                    row_count = 0
                    for row in reader:
                        cleaned_row = {}
                        for k, v in row.items():
                            key = self._normalize_header(k, norm_mode) if k else ""
                            if key:
                                cleaned_row[key] = v.strip() if v and self.config.trim_whitespace else (v or "")
                        rows.append(cleaned_row)
                        row_count += 1

                    self.stats.files_processed += 1
                    self.stats.total_rows_read += row_count
                    self.log(f"✓ {file_path.name} ({row_count:,} rows, {encoding}, delim={repr(detected_delim)})", "success")
                    return rows

            except UnicodeDecodeError:
                continue
            except Exception as e:
                if encoding == encodings_to_try[-1]:
                    self.log(f"✗ Error reading {file_path.name}: {e}", "error")
                    self.stats.files_skipped += 1
                    self.stats.errors.append(f"{file_path.name}: {e}")
                continue

        return rows
    
    def _get_final_columns(self, discovered_order: list) -> list[str]:
        """Determine final column list based on configuration."""
        if self.config.column_order:
            return self.config.column_order
        
        if self.config.columns_mode == "all":
            columns = discovered_order
        elif self.config.columns_mode == "select":
            columns = [c for c in discovered_order if c in self.config.selected_columns]
        elif self.config.columns_mode == "exclude":
            columns = [c for c in discovered_order if c not in self.config.selected_columns]
        else:
            columns = discovered_order
        
        # Apply column mapping for display names
        return columns

    def _with_transform_columns(self, columns: list[str]) -> list[str]:
        """Add output columns created by split, merge, and compute transforms."""
        result = list(columns)
        for ct in getattr(self.config, 'column_transforms', []):
            if len(ct) >= 3 and ct[1] == "split":
                num_parts = int(ct[3]) if len(ct) > 3 else 2
                for i in range(1, num_parts + 1):
                    new_col = f"{ct[0]}_{i}"
                    if new_col not in result:
                        result.append(new_col)
            elif len(ct) >= 4 and ct[1] == "merge":
                new_col = ct[0]
                if new_col not in result:
                    result.append(new_col)
            elif len(ct) >= 3 and ct[1] == "compute":
                new_col = ct[0]
                if new_col not in result:
                    result.append(new_col)
        return result
    
    def _apply_filters(self, rows: list[dict]) -> list[dict]:
        """Apply configured filters to rows."""
        if not self.config.filters:
            return rows
        
        filtered = []
        original_count = len(rows)
        
        for row in rows:
            if self._row_matches_filters(row):
                filtered.append(row)
        
        self.stats.rows_filtered = original_count - len(filtered)
        self.log(f"  Filtered out {self.stats.rows_filtered:,} rows", "info")
        
        return filtered

    def _row_matches_filters(self, row: dict) -> bool:
        results = []
        for col, operator, value in self.config.filters:
            cell_value = row.get(col, "")
            op_func = self.FILTER_OPERATORS.get(operator)
            if op_func:
                results.append(op_func(cell_value, value))
            else:
                results.append(True)

        if self.config.filter_logic == "and":
            return all(results) if results else True
        return any(results) if results else True
    
    def _apply_transformations(self, rows: list[dict], columns: list[str]) -> list[dict]:
        """Apply configured transformations to data."""
        # Pre-index per-column transforms for fast lookup
        col_transforms = {}  # {column: [(transform_type, *args), ...]}
        for ct in getattr(self.config, 'column_transforms', []):
            col_name = ct[0]
            transform = ct[1:]
            col_transforms.setdefault(col_name, []).append(transform)

        # Collect split/merge/compute transforms that alter the column set
        extra_columns = []
        for ct in getattr(self.config, 'column_transforms', []):
            if len(ct) >= 3 and ct[1] == "split":
                # split: (column, "split", delimiter, num_parts)
                num_parts = int(ct[3]) if len(ct) > 3 else 2
                for i in range(1, num_parts + 1):
                    new_col = f"{ct[0]}_{i}"
                    if new_col not in columns and new_col not in extra_columns:
                        extra_columns.append(new_col)
            elif len(ct) >= 4 and ct[1] == "merge":
                # merge: (target_col, "merge", separator, col1, col2, ...)
                new_col = ct[0]
                if new_col not in columns and new_col not in extra_columns:
                    extra_columns.append(new_col)
            elif len(ct) >= 3 and ct[1] == "compute":
                # compute: (new_column, "compute", expression)
                new_col = ct[0]
                if new_col not in columns and new_col not in extra_columns:
                    extra_columns.append(new_col)

        all_columns = columns + extra_columns

        transformed = []

        for row in rows:
            new_row = {}
            for col in columns:
                value = row.get(col, self.config.empty_value)

                if value is None:
                    value = self.config.empty_value

                # Global case transformation
                if self.config.case_transform == "upper":
                    value = str(value).upper()
                elif self.config.case_transform == "lower":
                    value = str(value).lower()
                elif self.config.case_transform == "title":
                    value = str(value).title()

                # Per-column transforms
                if col in col_transforms:
                    for transform in col_transforms[col]:
                        value = self._apply_column_transform(value, transform, row)

                # Apply column mapping
                output_col = self.config.column_mapping.get(col, col)
                new_row[output_col] = value

            # Handle split column transforms
            for ct in getattr(self.config, 'column_transforms', []):
                if len(ct) >= 3 and ct[1] == "split":
                    src_col = ct[0]
                    delimiter = ct[2]
                    num_parts = int(ct[3]) if len(ct) > 3 else 2
                    src_value = str(row.get(src_col, ""))
                    parts = src_value.split(delimiter, num_parts - 1)
                    for i in range(num_parts):
                        new_col = f"{src_col}_{i+1}"
                        new_row[new_col] = parts[i].strip() if i < len(parts) else ""

            # Handle merge column transforms
            for ct in getattr(self.config, 'column_transforms', []):
                if len(ct) >= 4 and ct[1] == "merge":
                    target_col = ct[0]
                    separator = ct[2]
                    merge_cols = ct[3:]
                    merged_parts = [str(row.get(c, new_row.get(c, ""))) for c in merge_cols]
                    new_row[target_col] = separator.join(merged_parts)

            # Handle compute column transforms
            for ct in getattr(self.config, 'column_transforms', []):
                if len(ct) >= 3 and ct[1] == "compute":
                    target_col = ct[0]
                    expression = ct[2]
                    new_row[target_col] = self._evaluate_expression(expression, row, new_row)

            transformed.append(new_row)

        return transformed

    @staticmethod
    def _apply_column_transform(value: str, transform: tuple, row: dict) -> str:
        """Apply a single per-column transform."""
        transform_type = transform[0]
        value = str(value)

        if transform_type == "trim":
            return value.strip()
        elif transform_type == "upper":
            return value.upper()
        elif transform_type == "lower":
            return value.lower()
        elif transform_type == "title":
            return value.title()
        elif transform_type == "replace" and len(transform) >= 3:
            # (replace, search, replacement)
            return value.replace(transform[1], transform[2])
        elif transform_type == "regex_replace" and len(transform) >= 3:
            # (regex_replace, pattern, replacement)
            try:
                return re.sub(transform[1], transform[2], value)
            except re.error:
                return value
        return value

    @staticmethod
    def _evaluate_expression(expression: str, row: dict, new_row: dict) -> str:
        """Evaluate a simple compute expression using row values.
        Supports: column references in curly braces, basic arithmetic.
        E.g. '{qty} * {price}' or '{first_name} + " " + {last_name}'
        """
        try:
            # Replace {column_name} with values
            resolved = expression
            # Find all column references
            col_refs = re.findall(r'\{([^}]+)\}', expression)
            for ref in col_refs:
                val = row.get(ref, new_row.get(ref, ""))
                # Try to use numeric value if possible
                try:
                    numeric_val = float(str(val).replace(",", ""))
                    resolved = resolved.replace(f'{{{ref}}}', str(numeric_val))
                except ValueError:
                    resolved = resolved.replace(f'{{{ref}}}', repr(str(val)))

            # Only allow safe operations
            allowed_chars = set('0123456789.+-*/() "\'_abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ,')
            if all(c in allowed_chars for c in resolved):
                result = eval(resolved)  # nosec - input is sanitized
                if isinstance(result, float) and result == int(result):
                    return str(int(result))
                return str(result)
            return ""
        except Exception:
            return ""
    
    def _deduplicate(self, rows: list[dict], columns: list[str]) -> list[dict]:
        """Remove duplicate rows."""
        if not rows:
            return rows
        
        dedupe_cols = self.config.dedupe_columns if self.config.dedupe_columns else columns
        fuzzy_enabled = getattr(self.config, "dedupe_fuzzy_enabled", False)
        fuzzy_threshold = int(getattr(self.config, "dedupe_fuzzy_threshold", 90))
        aggregate_mode = getattr(self.config, "dedupe_aggregate_mode", "none")

        seen = {}
        seen_keys = []
        result = []
        
        for row in rows:
            key = self._dedupe_key(row, dedupe_cols)

            matched_key = None
            if fuzzy_enabled:
                key_text = "\x1f".join(key)
                for existing_key in seen_keys:
                    existing_text = "\x1f".join(existing_key)
                    if self._fuzzy_score(key_text, existing_text) >= fuzzy_threshold:
                        matched_key = existing_key
                        break
            elif key in seen:
                matched_key = key

            if matched_key is None:
                seen[key] = len(result)
                seen_keys.append(key)
                result.append(row)
            elif aggregate_mode != "none":
                result[seen[matched_key]] = self._aggregate_duplicate_row(
                    result[seen[matched_key]],
                    row,
                    dedupe_cols,
                    columns,
                    aggregate_mode,
                )
            elif self.config.dedupe_keep == "last":
                result[seen[matched_key]] = row
        
        self.stats.duplicates_removed = len(rows) - len(result)
        self.log(f"  Removed {self.stats.duplicates_removed:,} duplicates", "info")
        
        return result

    def _dedupe_key(self, row: dict, dedupe_cols: list[str]) -> tuple[str, ...]:
        key_parts = []
        for col in dedupe_cols:
            mapped_col = self.config.column_mapping.get(col, col)
            val = row.get(mapped_col, row.get(col, ""))
            key_parts.append(str(val).lower() if not self.config.sort_case_sensitive else str(val))
        return tuple(key_parts)

    @staticmethod
    def _fuzzy_score(left: str, right: str) -> float:
        left_lower = left.lower()
        right_lower = right.lower()
        if left_lower in right_lower or right_lower in left_lower:
            return 100
        try:
            from rapidfuzz import fuzz
            return fuzz.partial_ratio(left_lower, right_lower)
        except ImportError:
            from difflib import SequenceMatcher
            return SequenceMatcher(None, left_lower, right_lower).ratio() * 100

    def _aggregate_duplicate_row(
        self,
        existing: dict,
        incoming: dict,
        dedupe_cols: list[str],
        output_columns: list[str],
        mode: str,
    ) -> dict:
        result = dict(existing)
        key_columns = {
            self.config.column_mapping.get(col, col)
            for col in dedupe_cols
        } | set(dedupe_cols)
        separator = getattr(self.config, "dedupe_aggregate_separator", "; ")

        for col in output_columns:
            mapped_col = self.config.column_mapping.get(col, col)
            if col in key_columns or mapped_col in key_columns:
                continue

            left = result.get(mapped_col, result.get(col, ""))
            right = incoming.get(mapped_col, incoming.get(col, ""))
            result[mapped_col] = self._aggregate_value(left, right, mode, separator)

        return result

    @staticmethod
    def _aggregate_value(left, right, mode: str, separator: str):
        left_text = "" if left is None else str(left)
        right_text = "" if right is None else str(right)
        if mode == "concat":
            parts = []
            for value in (left_text, right_text):
                if value and value not in parts:
                    parts.append(value)
            return separator.join(parts)

        left_num = CSVEngine._parse_number(left_text)
        right_num = CSVEngine._parse_number(right_text)
        if left_num is None or right_num is None:
            if mode == "max":
                return max(left_text, right_text)
            if mode == "min":
                return min(left_text, right_text)
            return left_text or right_text

        if mode == "sum":
            value = left_num + right_num
        elif mode == "max":
            value = max(left_num, right_num)
        elif mode == "min":
            value = min(left_num, right_num)
        else:
            return left_text

        return str(int(value)) if value == int(value) else str(value)
    
    def _sort_rows(self, rows: list[dict], columns: list[str]) -> list[dict]:
        """Sort rows by configured columns."""
        if not rows or not self.config.sort_columns:
            return rows
        
        def sort_key(row):
            keys = []
            for col, ascending in self.config.sort_columns:
                mapped_col = self.config.column_mapping.get(col, col)
                val = row.get(mapped_col, row.get(col, ""))
                
                if self.config.sort_numeric_aware:
                    try:
                        val = float(str(val).replace(",", ""))
                    except ValueError:
                        if not self.config.sort_case_sensitive:
                            val = str(val).lower()
                else:
                    if not self.config.sort_case_sensitive:
                        val = str(val).lower()
                
                keys.append(val)
            return keys
        
        # Determine sort direction for each column
        reverse_flags = [not asc for _, asc in self.config.sort_columns]
        
        # Multi-column sort requires custom approach
        # Sort by last column first, then work backwards
        sorted_rows = rows.copy()
        for i in range(len(self.config.sort_columns) - 1, -1, -1):
            col, ascending = self.config.sort_columns[i]
            mapped_col = self.config.column_mapping.get(col, col)
            
            def make_key(row, mc=mapped_col, c=col, na=self.config.sort_numeric_aware, cs=self.config.sort_case_sensitive):
                val = row.get(mc, row.get(c, ""))
                if na:
                    parsed_number = CSVEngine._parse_number(val)
                    if parsed_number is not None:
                        return (0, parsed_number)
                    parsed_date = CSVEngine._parse_datetime(val)
                    if parsed_date is not None:
                        return (1, parsed_date.timestamp())
                text = str(val) if cs else str(val).lower()
                return (2, locale.strxfrm(text))
            
            sorted_rows.sort(key=make_key, reverse=not ascending)
        
        self.log(f"  Sorted by {len(self.config.sort_columns)} column(s)", "info")
        
        return sorted_rows
    
    def _write_output(self, rows: list[dict], columns: list[str], output_file: Path):
        """Write processed data to output file."""
        if not rows:
            self.log("No data to write", "warning")
            return
        
        # Map columns to output names
        output_columns = [self.config.column_mapping.get(c, c) for c in columns]

        suffix = output_file.suffix.lower()
        if suffix == ".xlsx":
            self._write_xlsx_output(rows, output_columns, output_file)
            return
        if suffix == ".parquet":
            self._write_parquet_output(rows, output_columns, output_file)
            return
        
        quoting_map = {
            "minimal": csv.QUOTE_MINIMAL,
            "all": csv.QUOTE_ALL,
            "nonnumeric": csv.QUOTE_NONNUMERIC,
            "none": csv.QUOTE_NONE,
        }
        quoting = quoting_map.get(self.config.output_quoting, csv.QUOTE_MINIMAL)
        
        # Line ending
        if self.config.line_ending == "unix":
            newline = "\n"
        elif self.config.line_ending == "windows":
            newline = "\r\n"
        else:
            newline = ""
        
        try:
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_file, 'w', encoding=self.config.output_encoding, 
                      newline=newline if newline else '') as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=output_columns,
                    delimiter=self.config.output_delimiter,
                    quoting=quoting,
                    extrasaction='ignore'
                )
                
                if self.config.include_header:
                    writer.writeheader()
                
                for row in rows:
                    writer.writerow(row)
            
            self.log(f"✓ Saved: {output_file.name} ({len(rows):,} rows)", "success")
            
        except Exception as e:
            self.log(f"✗ Error writing file: {e}", "error")
            self.stats.errors.append(f"Write error: {e}")

    def _write_xlsx_output(self, rows: list[dict], output_columns: list[str], output_file: Path):
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            self.log("XX openpyxl is required for .xlsx output", "error")
            self.stats.errors.append(f"Write error: {exc}")
            return

        try:
            output_file.parent.mkdir(parents=True, exist_ok=True)
            workbook = Workbook(write_only=True)
            sheet = workbook.create_sheet("Data")
            if self.config.include_header:
                sheet.append(output_columns)
            for row in rows:
                sheet.append([row.get(column, "") for column in output_columns])
            workbook.save(output_file)
            self.log(f"OK Saved: {output_file.name} ({len(rows):,} rows)", "success")
        except Exception as exc:
            self.log(f"XX Error writing file: {exc}", "error")
            self.stats.errors.append(f"Write error: {exc}")

    def _write_parquet_output(self, rows: list[dict], output_columns: list[str], output_file: Path):
        try:
            output_file.parent.mkdir(parents=True, exist_ok=True)
            records = [
                {column: row.get(column, "") for column in output_columns}
                for row in rows
            ]
            try:
                import pyarrow as pa
                import pyarrow.parquet as pq
                table = pa.Table.from_pylist(records)
                pq.write_table(table, output_file)
            except ImportError:
                import polars as pl
                pl.DataFrame(records).write_parquet(output_file)
            self.log(f"OK Saved: {output_file.name} ({len(rows):,} rows)", "success")
        except Exception as exc:
            self.log(f"XX Error writing file: {exc}", "error")
            self.stats.errors.append(f"Write error: {exc}")


# ══════════════════════════════════════════════════════════════════════════════
# GUI COMPONENTS
# ══════════════════════════════════════════════════════════════════════════════

class FileListPanel(ctk.CTkFrame):
    """File list with drag-drop support."""
    
    def __init__(self, master, on_change: Callable = None, **kwargs):
        if "fg_color" not in kwargs:
            kwargs["fg_color"] = COLORS["bg_secondary"]
        if "corner_radius" not in kwargs:
            kwargs["corner_radius"] = 8
        super().__init__(master, **kwargs)
        
        self.files: list[Path] = []
        self.on_change = on_change
        
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=12, pady=(12, 8))
        
        ctk.CTkLabel(
            header, text="📁 Input Files",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left")
        
        self.count_label = ctk.CTkLabel(
            header, text="0 files",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_muted"]
        )
        self.count_label.pack(side="right")
        
        # List container
        list_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_dark"], corner_radius=6)
        list_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        
        self.scroll_frame = ctk.CTkScrollableFrame(
            list_frame, fg_color="transparent",
            scrollbar_button_color=COLORS["bg_tertiary"],
            scrollbar_button_hover_color=COLORS["accent_blue"]
        )
        self.scroll_frame.pack(fill="both", expand=True, padx=4, pady=4)
        
        self.placeholder = ctk.CTkLabel(
            self.scroll_frame,
            text="Drag & drop CSV files here\nor use buttons below",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_muted"]
        )
        self.placeholder.pack(expand=True, pady=30)
        
        # Buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=12, pady=(0, 12))
        
        ctk.CTkButton(
            btn_frame, text="Add Files", font=ctk.CTkFont(size=12),
            height=32, fg_color=COLORS["accent_blue"],
            hover_color=COLORS["accent_blue_hover"],
            corner_radius=6, command=self._browse_files
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))
        
        ctk.CTkButton(
            btn_frame, text="Add Folder", font=ctk.CTkFont(size=12),
            height=32, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_hover"],
            text_color=COLORS["text_primary"],
            corner_radius=6, command=self._browse_folder
        ).pack(side="left", fill="x", expand=True, padx=(4, 4))
        
        ctk.CTkButton(
            btn_frame, text="Clear", font=ctk.CTkFont(size=12),
            height=32, width=60, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["accent_red"],
            text_color=COLORS["text_secondary"],
            corner_radius=6, command=self.clear
        ).pack(side="left", padx=(4, 0))
    
    def _browse_files(self):
        files = filedialog.askopenfilenames(
            title="Select Input Files",
            filetypes=[
                ("Supported Files", "*.csv *.tsv *.txt *.xlsx *.parquet"),
                ("CSV Files", "*.csv"),
                ("TSV Files", "*.tsv"),
                ("Excel Files", "*.xlsx"),
                ("Parquet Files", "*.parquet"),
                ("Text Files", "*.txt"),
                ("All Files", "*.*"),
            ]
        )
        if files:
            self.add_files([Path(f) for f in files])
    
    def _browse_folder(self):
        folder = filedialog.askdirectory(title="Select Folder")
        if folder:
            folder_path = Path(folder)
            input_files = [
                path for path in folder_path.rglob("*")
                if path.is_file() and path.suffix.lower() in SUPPORTED_INPUT_SUFFIXES
            ]
            if input_files:
                self.add_files(input_files)
    
    def add_files(self, paths: list[Path]) -> int:
        added = 0
        for p in paths:
            path = Path(p) if not isinstance(p, Path) else p
            if path.is_dir():
                added += self.add_files([
                    child for child in path.rglob("*")
                    if child.is_file() and child.suffix.lower() in SUPPORTED_INPUT_SUFFIXES
                ])
            elif path.suffix.lower() in SUPPORTED_INPUT_SUFFIXES and path not in self.files:
                self.files.append(path)
                added += 1
        
        if added:
            self._refresh()
            if self.on_change:
                self.on_change()
        return added
    
    def remove_file(self, path: Path):
        if path in self.files:
            self.files.remove(path)
            self._refresh()
            if self.on_change:
                self.on_change()
    
    def clear(self):
        self.files.clear()
        self._refresh()
        if self.on_change:
            self.on_change()
    
    def _refresh(self):
        for w in self.scroll_frame.winfo_children():
            w.destroy()
        
        if not self.files:
            self.placeholder = ctk.CTkLabel(
                self.scroll_frame,
                text="Drag & drop CSV files here\nor use buttons below",
                font=ctk.CTkFont(size=12),
                text_color=COLORS["text_muted"]
            )
            self.placeholder.pack(expand=True, pady=30)
            self.count_label.configure(text="0 files")
        else:
            for idx, fp in enumerate(self.files):
                self._create_item(fp, idx)
            self.count_label.configure(text=f"{len(self.files)} file{'s' if len(self.files) != 1 else ''}")
    
    def _create_item(self, path: Path, idx: int):
        bg = COLORS["bg_secondary"] if idx % 2 == 0 else COLORS["bg_tertiary"]
        
        frame = ctk.CTkFrame(self.scroll_frame, fg_color=bg, corner_radius=4, height=32)
        frame.pack(fill="x", pady=1)
        frame.pack_propagate(False)
        
        ctk.CTkLabel(frame, text="📄", font=ctk.CTkFont(size=12), width=24).pack(side="left", padx=(6, 2))
        
        ctk.CTkLabel(
            frame, text=path.name, font=ctk.CTkFont(size=11),
            text_color=COLORS["text_primary"], anchor="w"
        ).pack(side="left", fill="x", expand=True)
        
        try:
            size = path.stat().st_size
            size_text = f"{size/1024:.1f}KB" if size >= 1024 else f"{size}B"
        except:
            size_text = ""
        
        ctk.CTkLabel(
            frame, text=size_text, font=ctk.CTkFont(size=10),
            text_color=COLORS["text_muted"], width=50
        ).pack(side="right", padx=4)
        
        ctk.CTkButton(
            frame, text="✕", font=ctk.CTkFont(size=10),
            width=24, height=24, fg_color="transparent",
            hover_color=COLORS["accent_red"],
            text_color=COLORS["text_muted"], corner_radius=4,
            command=lambda p=path: self.remove_file(p)
        ).pack(side="right", padx=2)


class ColumnPanel(ctk.CTkFrame):
    """Column selection and configuration."""
    
    def __init__(self, master, **kwargs):
        if "fg_color" not in kwargs:
            kwargs["fg_color"] = COLORS["bg_secondary"]
        if "corner_radius" not in kwargs:
            kwargs["corner_radius"] = 8
        super().__init__(master, **kwargs)
        
        self.columns: list[str] = []
        self.selected: set[str] = set()
        self.column_mapping: dict[str, str] = {}
        self.mode = StringVar(value="all")
        
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=12, pady=(12, 8))
        
        ctk.CTkLabel(
            header, text="📊 Column Selection",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left")
        
        self.count_label = ctk.CTkLabel(
            header, text="No columns",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_muted"]
        )
        self.count_label.pack(side="right")
        
        # Mode selection
        mode_frame = ctk.CTkFrame(self, fg_color="transparent")
        mode_frame.pack(fill="x", padx=12, pady=(0, 8))
        
        for mode, text in [("all", "All Columns"), ("select", "Include Selected"), ("exclude", "Exclude Selected")]:
            ctk.CTkRadioButton(
                mode_frame, text=text, variable=self.mode, value=mode,
                font=ctk.CTkFont(size=11),
                fg_color=COLORS["accent_blue"],
                hover_color=COLORS["accent_blue_hover"],
                text_color=COLORS["text_secondary"]
            ).pack(side="left", padx=(0, 12))
        
        # Column list
        list_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_dark"], corner_radius=6)
        list_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        
        self.scroll_frame = ctk.CTkScrollableFrame(
            list_frame, fg_color="transparent",
            scrollbar_button_color=COLORS["bg_tertiary"],
            scrollbar_button_hover_color=COLORS["accent_blue"]
        )
        self.scroll_frame.pack(fill="both", expand=True, padx=4, pady=4)
        
        self.placeholder = ctk.CTkLabel(
            self.scroll_frame,
            text="Add files to discover columns",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_muted"]
        )
        self.placeholder.pack(expand=True, pady=20)
        
        # Buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=12, pady=(0, 12))
        
        ctk.CTkButton(
            btn_frame, text="Select All", font=ctk.CTkFont(size=11),
            height=28, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=4, command=self._select_all
        ).pack(side="left", padx=(0, 4))
        
        ctk.CTkButton(
            btn_frame, text="Select None", font=ctk.CTkFont(size=11),
            height=28, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=4, command=self._select_none
        ).pack(side="left")
    
    def set_columns(self, columns: list[str]):
        self.columns = columns
        self.selected = set(columns)
        self._refresh()
    
    def _select_all(self):
        self.selected = set(self.columns)
        self._refresh()
    
    def _select_none(self):
        self.selected.clear()
        self._refresh()
    
    def _toggle_column(self, col: str, var: BooleanVar):
        if var.get():
            self.selected.add(col)
        else:
            self.selected.discard(col)
    
    def _refresh(self):
        for w in self.scroll_frame.winfo_children():
            w.destroy()
        
        if not self.columns:
            self.placeholder = ctk.CTkLabel(
                self.scroll_frame,
                text="Add files to discover columns",
                font=ctk.CTkFont(size=12),
                text_color=COLORS["text_muted"]
            )
            self.placeholder.pack(expand=True, pady=20)
            self.count_label.configure(text="No columns")
        else:
            for col in self.columns:
                var = BooleanVar(value=col in self.selected)
                
                frame = ctk.CTkFrame(self.scroll_frame, fg_color="transparent", height=28)
                frame.pack(fill="x", pady=1)
                
                ctk.CTkCheckBox(
                    frame, text=col, variable=var,
                    font=ctk.CTkFont(size=11),
                    fg_color=COLORS["accent_blue"],
                    hover_color=COLORS["accent_blue_hover"],
                    text_color=COLORS["text_primary"],
                    command=lambda c=col, v=var: self._toggle_column(c, v)
                ).pack(side="left")
            
            self.count_label.configure(text=f"{len(self.columns)} columns")
    
    def get_config(self) -> tuple[str, list[str]]:
        return self.mode.get(), list(self.selected)


class SortPanel(ctk.CTkFrame):
    """Sorting configuration."""
    
    def __init__(self, master, **kwargs):
        if "fg_color" not in kwargs:
            kwargs["fg_color"] = COLORS["bg_secondary"]
        if "corner_radius" not in kwargs:
            kwargs["corner_radius"] = 8
        super().__init__(master, **kwargs)
        
        self.columns: list[str] = []
        self.sort_rules: list[tuple[str, bool]] = []  # (column, ascending)
        
        self.enabled = BooleanVar(value=False)
        self.case_sensitive = BooleanVar(value=False)
        self.numeric_aware = BooleanVar(value=True)
        
        # Header with enable toggle
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=12, pady=(12, 8))
        
        ctk.CTkSwitch(
            header, text="🔤 Sorting",
            variable=self.enabled,
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_primary"],
            fg_color=COLORS["bg_tertiary"],
            progress_color=COLORS["accent_green"],
            button_color=COLORS["text_secondary"],
            button_hover_color=COLORS["text_primary"]
        ).pack(side="left")
        
        # Options
        opts_frame = ctk.CTkFrame(self, fg_color="transparent")
        opts_frame.pack(fill="x", padx=12, pady=(0, 8))
        
        ctk.CTkCheckBox(
            opts_frame, text="Case sensitive", variable=self.case_sensitive,
            font=ctk.CTkFont(size=11),
            fg_color=COLORS["accent_blue"],
            text_color=COLORS["text_secondary"]
        ).pack(side="left", padx=(0, 12))
        
        ctk.CTkCheckBox(
            opts_frame, text="Numeric-aware", variable=self.numeric_aware,
            font=ctk.CTkFont(size=11),
            fg_color=COLORS["accent_blue"],
            text_color=COLORS["text_secondary"]
        ).pack(side="left")
        
        # Sort rules list
        list_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_dark"], corner_radius=6)
        list_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        
        self.rules_frame = ctk.CTkScrollableFrame(
            list_frame, fg_color="transparent",
            scrollbar_button_color=COLORS["bg_tertiary"]
        )
        self.rules_frame.pack(fill="both", expand=True, padx=4, pady=4)
        
        # Add rule button
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=12, pady=(0, 12))
        
        ctk.CTkButton(
            btn_frame, text="+ Add Sort Rule", font=ctk.CTkFont(size=11),
            height=28, fg_color=COLORS["accent_purple"],
            hover_color=COLORS["accent_purple_hover"],
            corner_radius=4, command=self._add_rule
        ).pack(side="left")
        
        ctk.CTkButton(
            btn_frame, text="Clear All", font=ctk.CTkFont(size=11),
            height=28, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["accent_red"],
            text_color=COLORS["text_secondary"],
            corner_radius=4, command=self._clear_rules
        ).pack(side="right")
    
    def set_columns(self, columns: list[str]):
        self.columns = columns
    
    def _add_rule(self):
        if not self.columns:
            return
        
        self.sort_rules.append((self.columns[0], True))
        self._refresh_rules()
    
    def _remove_rule(self, idx: int):
        if 0 <= idx < len(self.sort_rules):
            self.sort_rules.pop(idx)
            self._refresh_rules()
    
    def _clear_rules(self):
        self.sort_rules.clear()
        self._refresh_rules()
    
    def _update_rule(self, idx: int, column: str = None, ascending: bool = None):
        if 0 <= idx < len(self.sort_rules):
            col, asc = self.sort_rules[idx]
            self.sort_rules[idx] = (column if column is not None else col,
                                     ascending if ascending is not None else asc)
    
    def _refresh_rules(self):
        for w in self.rules_frame.winfo_children():
            w.destroy()
        
        if not self.sort_rules:
            ctk.CTkLabel(
                self.rules_frame,
                text="No sort rules defined",
                font=ctk.CTkFont(size=11),
                text_color=COLORS["text_muted"]
            ).pack(pady=10)
        else:
            for idx, (col, asc) in enumerate(self.sort_rules):
                self._create_rule_row(idx, col, asc)
    
    def _create_rule_row(self, idx: int, column: str, ascending: bool):
        frame = ctk.CTkFrame(self.rules_frame, fg_color=COLORS["bg_secondary"], corner_radius=4, height=36)
        frame.pack(fill="x", pady=2)
        frame.pack_propagate(False)
        
        ctk.CTkLabel(
            frame, text=f"{idx + 1}.",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_muted"], width=24
        ).pack(side="left", padx=(8, 4))
        
        col_var = StringVar(value=column)
        col_menu = ctk.CTkOptionMenu(
            frame, variable=col_var, values=self.columns,
            font=ctk.CTkFont(size=11), height=28, width=150,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            button_hover_color=COLORS["bg_hover"],
            dropdown_fg_color=COLORS["bg_secondary"],
            command=lambda v, i=idx: self._update_rule(i, column=v)
        )
        col_menu.pack(side="left", padx=4)
        
        dir_var = StringVar(value="A→Z" if ascending else "Z→A")
        dir_menu = ctk.CTkOptionMenu(
            frame, variable=dir_var, values=["A→Z", "Z→A"],
            font=ctk.CTkFont(size=11), height=28, width=70,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            button_hover_color=COLORS["bg_hover"],
            dropdown_fg_color=COLORS["bg_secondary"],
            command=lambda v, i=idx: self._update_rule(i, ascending=(v == "A→Z"))
        )
        dir_menu.pack(side="left", padx=4)
        
        ctk.CTkButton(
            frame, text="✕", font=ctk.CTkFont(size=10),
            width=24, height=24, fg_color="transparent",
            hover_color=COLORS["accent_red"],
            text_color=COLORS["text_muted"], corner_radius=4,
            command=lambda i=idx: self._remove_rule(i)
        ).pack(side="right", padx=4)
    
    def get_config(self) -> tuple[bool, list, bool, bool]:
        return self.enabled.get(), self.sort_rules.copy(), self.case_sensitive.get(), self.numeric_aware.get()


class DedupePanel(ctk.CTkFrame):
    """Deduplication configuration."""
    
    def __init__(self, master, **kwargs):
        if "fg_color" not in kwargs:
            kwargs["fg_color"] = COLORS["bg_secondary"]
        if "corner_radius" not in kwargs:
            kwargs["corner_radius"] = 8
        super().__init__(master, **kwargs)
        
        self.columns: list[str] = []
        self.selected_columns: set[str] = set()
        
        self.enabled = BooleanVar(value=True)
        self.keep_mode = StringVar(value="first")
        self.use_all_columns = BooleanVar(value=True)
        self.fuzzy_enabled = BooleanVar(value=False)
        self.fuzzy_threshold = IntVar(value=90)
        self.aggregate_mode = StringVar(value="none")
        self.aggregate_separator = StringVar(value="; ")
        
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=12, pady=(12, 8))
        
        ctk.CTkSwitch(
            header, text="🔄 Deduplication",
            variable=self.enabled,
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_primary"],
            fg_color=COLORS["bg_tertiary"],
            progress_color=COLORS["accent_green"]
        ).pack(side="left")
        
        # Keep mode
        keep_frame = ctk.CTkFrame(self, fg_color="transparent")
        keep_frame.pack(fill="x", padx=12, pady=(0, 8))
        
        ctk.CTkLabel(
            keep_frame, text="Keep:",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"]
        ).pack(side="left", padx=(0, 8))
        
        for val, text in [("first", "First"), ("last", "Last")]:
            ctk.CTkRadioButton(
                keep_frame, text=text, variable=self.keep_mode, value=val,
                font=ctk.CTkFont(size=11),
                fg_color=COLORS["accent_blue"],
                text_color=COLORS["text_secondary"]
            ).pack(side="left", padx=(0, 12))
        
        # Column selection mode
        col_mode_frame = ctk.CTkFrame(self, fg_color="transparent")
        col_mode_frame.pack(fill="x", padx=12, pady=(0, 8))
        
        ctk.CTkCheckBox(
            col_mode_frame, text="Use all columns for comparison",
            variable=self.use_all_columns,
            font=ctk.CTkFont(size=11),
            fg_color=COLORS["accent_blue"],
            text_color=COLORS["text_secondary"],
            command=self._toggle_column_selection
        ).pack(side="left")

        fuzzy_frame = ctk.CTkFrame(self, fg_color="transparent")
        fuzzy_frame.pack(fill="x", padx=12, pady=(0, 8))

        ctk.CTkCheckBox(
            fuzzy_frame, text="Fuzzy duplicate matching",
            variable=self.fuzzy_enabled,
            font=ctk.CTkFont(size=11),
            fg_color=COLORS["accent_blue"],
            text_color=COLORS["text_secondary"],
        ).pack(side="left", padx=(0, 12))

        ctk.CTkLabel(
            fuzzy_frame, text="Threshold",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_muted"],
        ).pack(side="left", padx=(0, 6))

        ctk.CTkSlider(
            fuzzy_frame, from_=50, to=100, number_of_steps=50,
            variable=self.fuzzy_threshold,
            width=100,
            progress_color=COLORS["accent_blue"],
            button_color=COLORS["accent_blue"],
            button_hover_color=COLORS["accent_blue_hover"],
        ).pack(side="left")

        aggregate_frame = ctk.CTkFrame(self, fg_color="transparent")
        aggregate_frame.pack(fill="x", padx=12, pady=(0, 8))

        ctk.CTkLabel(
            aggregate_frame, text="Aggregate:",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"],
        ).pack(side="left", padx=(0, 8))

        ctk.CTkOptionMenu(
            aggregate_frame, variable=self.aggregate_mode,
            values=["none", "max", "min", "sum", "concat"],
            font=ctk.CTkFont(size=11), height=28, width=90,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            dropdown_fg_color=COLORS["bg_secondary"],
        ).pack(side="left", padx=(0, 8))

        ctk.CTkEntry(
            aggregate_frame, textvariable=self.aggregate_separator,
            placeholder_text="separator",
            font=ctk.CTkFont(size=11), height=28, width=80,
            fg_color=COLORS["bg_dark"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"],
        ).pack(side="left")
        
        # Column list (hidden when use_all_columns is True)
        self.col_list_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_dark"], corner_radius=6)
        
        self.col_scroll = ctk.CTkScrollableFrame(
            self.col_list_frame, fg_color="transparent", height=100,
            scrollbar_button_color=COLORS["bg_tertiary"]
        )
        self.col_scroll.pack(fill="both", expand=True, padx=4, pady=4)
    
    def _toggle_column_selection(self):
        if self.use_all_columns.get():
            self.col_list_frame.pack_forget()
        else:
            self.col_list_frame.pack(fill="x", padx=12, pady=(0, 12))
            self._refresh_columns()
    
    def set_columns(self, columns: list[str]):
        self.columns = columns
        self.selected_columns = set(columns)
        self._refresh_columns()
    
    def _toggle_column(self, col: str, var: BooleanVar):
        if var.get():
            self.selected_columns.add(col)
        else:
            self.selected_columns.discard(col)
    
    def _refresh_columns(self):
        for w in self.col_scroll.winfo_children():
            w.destroy()
        
        for col in self.columns:
            var = BooleanVar(value=col in self.selected_columns)
            ctk.CTkCheckBox(
                self.col_scroll, text=col, variable=var,
                font=ctk.CTkFont(size=11),
                fg_color=COLORS["accent_blue"],
                text_color=COLORS["text_primary"],
                command=lambda c=col, v=var: self._toggle_column(c, v)
            ).pack(anchor="w", pady=1)
    
    def get_config(self) -> tuple[bool, list, str, bool, int, str, str]:
        columns = [] if self.use_all_columns.get() else list(self.selected_columns)
        return (
            self.enabled.get(),
            columns,
            self.keep_mode.get(),
            self.fuzzy_enabled.get(),
            int(self.fuzzy_threshold.get()),
            self.aggregate_mode.get(),
            self.aggregate_separator.get(),
        )


class FilterPanel(ctk.CTkFrame):
    """Filter configuration."""

    OPERATORS = [
        ("equals", "Equals"),
        ("not_equals", "Not Equals"),
        ("contains", "Contains"),
        ("not_contains", "Not Contains"),
        ("starts_with", "Starts With"),
        ("ends_with", "Ends With"),
        ("is_empty", "Is Empty"),
        ("is_not_empty", "Is Not Empty"),
        ("greater_than", "Greater Than"),
        ("less_than", "Less Than"),
        ("greater_than_or_equal", "Greater/Equal"),
        ("less_than_or_equal", "Less/Equal"),
        ("between", "Between"),
        ("fuzzy", "Fuzzy Match"),
        ("in_list", "In List"),
        ("not_in_list", "Not In List"),
        ("regex", "Regex Match"),
    ]
    
    def __init__(self, master, **kwargs):
        if "fg_color" not in kwargs:
            kwargs["fg_color"] = COLORS["bg_secondary"]
        if "corner_radius" not in kwargs:
            kwargs["corner_radius"] = 8
        super().__init__(master, **kwargs)
        
        self.columns: list[str] = []
        self.filters: list[tuple[str, str, str]] = []  # (column, operator, value)
        self.logic = StringVar(value="and")
        
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=12, pady=(12, 8))
        
        ctk.CTkLabel(
            header, text="🔍 Filters",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left")
        
        # Logic selection
        logic_frame = ctk.CTkFrame(header, fg_color="transparent")
        logic_frame.pack(side="right")
        
        ctk.CTkRadioButton(
            logic_frame, text="AND", variable=self.logic, value="and",
            font=ctk.CTkFont(size=10),
            fg_color=COLORS["accent_blue"],
            text_color=COLORS["text_secondary"]
        ).pack(side="left", padx=4)
        
        ctk.CTkRadioButton(
            logic_frame, text="OR", variable=self.logic, value="or",
            font=ctk.CTkFont(size=10),
            fg_color=COLORS["accent_blue"],
            text_color=COLORS["text_secondary"]
        ).pack(side="left")
        
        # Filter list
        list_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_dark"], corner_radius=6)
        list_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        
        self.filters_frame = ctk.CTkScrollableFrame(
            list_frame, fg_color="transparent",
            scrollbar_button_color=COLORS["bg_tertiary"]
        )
        self.filters_frame.pack(fill="both", expand=True, padx=4, pady=4)
        
        self.placeholder = ctk.CTkLabel(
            self.filters_frame,
            text="No filters defined",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_muted"]
        )
        self.placeholder.pack(pady=10)
        
        # Buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=12, pady=(0, 12))
        
        ctk.CTkButton(
            btn_frame, text="+ Add Filter", font=ctk.CTkFont(size=11),
            height=28, fg_color=COLORS["accent_orange"],
            hover_color="#ea580c",
            corner_radius=4, command=self._add_filter
        ).pack(side="left")
        
        ctk.CTkButton(
            btn_frame, text="Clear All", font=ctk.CTkFont(size=11),
            height=28, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["accent_red"],
            text_color=COLORS["text_secondary"],
            corner_radius=4, command=self._clear_filters
        ).pack(side="right")
    
    def set_columns(self, columns: list[str]):
        self.columns = columns
    
    def _add_filter(self):
        if not self.columns:
            return
        self.filters.append((self.columns[0], "contains", ""))
        self._refresh()
    
    def _remove_filter(self, idx: int):
        if 0 <= idx < len(self.filters):
            self.filters.pop(idx)
            self._refresh()
    
    def _clear_filters(self):
        self.filters.clear()
        self._refresh()
    
    def _update_filter(self, idx: int, column: str = None, operator: str = None, value: str = None):
        if 0 <= idx < len(self.filters):
            col, op, val = self.filters[idx]
            self.filters[idx] = (
                column if column is not None else col,
                operator if operator is not None else op,
                value if value is not None else val
            )
    
    def _refresh(self):
        for w in self.filters_frame.winfo_children():
            w.destroy()
        
        if not self.filters:
            self.placeholder = ctk.CTkLabel(
                self.filters_frame,
                text="No filters defined",
                font=ctk.CTkFont(size=11),
                text_color=COLORS["text_muted"]
            )
            self.placeholder.pack(pady=10)
        else:
            for idx, (col, op, val) in enumerate(self.filters):
                self._create_filter_row(idx, col, op, val)
    
    def _create_filter_row(self, idx: int, column: str, operator: str, value: str):
        frame = ctk.CTkFrame(self.filters_frame, fg_color=COLORS["bg_secondary"], corner_radius=4)
        frame.pack(fill="x", pady=2)
        
        row1 = ctk.CTkFrame(frame, fg_color="transparent")
        row1.pack(fill="x", padx=8, pady=(6, 2))
        
        col_var = StringVar(value=column)
        ctk.CTkOptionMenu(
            row1, variable=col_var, values=self.columns,
            font=ctk.CTkFont(size=10), height=26, width=120,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            dropdown_fg_color=COLORS["bg_secondary"],
            command=lambda v, i=idx: self._update_filter(i, column=v)
        ).pack(side="left", padx=(0, 4))
        
        op_display = {op: name for op, name in self.OPERATORS}
        op_var = StringVar(value=op_display.get(operator, operator))
        ctk.CTkOptionMenu(
            row1, variable=op_var, values=[name for _, name in self.OPERATORS],
            font=ctk.CTkFont(size=10), height=26, width=100,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            dropdown_fg_color=COLORS["bg_secondary"],
            command=lambda v, i=idx: self._update_filter(i, operator=next((op for op, name in self.OPERATORS if name == v), v))
        ).pack(side="left", padx=(0, 4))
        
        ctk.CTkButton(
            row1, text="✕", font=ctk.CTkFont(size=10),
            width=24, height=24, fg_color="transparent",
            hover_color=COLORS["accent_red"],
            text_color=COLORS["text_muted"], corner_radius=4,
            command=lambda i=idx: self._remove_filter(i)
        ).pack(side="right")
        
        # Value entry (hidden for is_empty/is_not_empty)
        if operator not in ["is_empty", "is_not_empty"]:
            row2 = ctk.CTkFrame(frame, fg_color="transparent")
            row2.pack(fill="x", padx=8, pady=(2, 6))
            
            val_entry = ctk.CTkEntry(
                row2, placeholder_text="Value...",
                font=ctk.CTkFont(size=10), height=26,
                fg_color=COLORS["bg_dark"],
                border_color=COLORS["border"],
                text_color=COLORS["text_primary"]
            )
            val_entry.pack(fill="x")
            val_entry.insert(0, value)
            val_entry.bind("<KeyRelease>", lambda e, i=idx: self._update_filter(i, value=e.widget.get()))
    
    def get_config(self) -> tuple[list, str]:
        return self.filters.copy(), self.logic.get()


class TransformPanel(ctk.CTkFrame):
    """Data transformation configuration."""

    COLUMN_TRANSFORM_TYPES = [
        ("trim", "Trim"),
        ("upper", "UPPER"),
        ("lower", "lower"),
        ("title", "Title"),
        ("replace", "Replace"),
        ("regex_replace", "Regex Replace"),
    ]

    def __init__(self, master, **kwargs):
        if "fg_color" not in kwargs:
            kwargs["fg_color"] = COLORS["bg_secondary"]
        if "corner_radius" not in kwargs:
            kwargs["corner_radius"] = 8
        super().__init__(master, **kwargs)

        self.columns: list[str] = []
        self.column_transforms: list = []  # [(column, type, *args)]

        self.trim_whitespace = BooleanVar(value=True)
        self.case_transform = StringVar(value="none")
        self.empty_value = StringVar(value="")
        self.header_normalize = StringVar(value="none")

        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=12, pady=(12, 8))

        ctk.CTkLabel(
            header, text="⚙️ Transformations",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left")

        # Options
        opts_frame = ctk.CTkFrame(self, fg_color="transparent")
        opts_frame.pack(fill="x", padx=12, pady=(0, 4))

        # Trim whitespace
        ctk.CTkCheckBox(
            opts_frame, text="Trim whitespace",
            variable=self.trim_whitespace,
            font=ctk.CTkFont(size=11),
            fg_color=COLORS["accent_blue"],
            text_color=COLORS["text_secondary"]
        ).pack(anchor="w", pady=4)

        # Case transformation
        case_frame = ctk.CTkFrame(opts_frame, fg_color="transparent")
        case_frame.pack(fill="x", pady=4)

        ctk.CTkLabel(
            case_frame, text="Case:",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"]
        ).pack(side="left", padx=(0, 8))

        for val, text in [("none", "None"), ("upper", "UPPER"), ("lower", "lower"), ("title", "Title")]:
            ctk.CTkRadioButton(
                case_frame, text=text, variable=self.case_transform, value=val,
                font=ctk.CTkFont(size=11),
                fg_color=COLORS["accent_blue"],
                text_color=COLORS["text_secondary"]
            ).pack(side="left", padx=(0, 8))

        # Header normalization
        header_frame = ctk.CTkFrame(opts_frame, fg_color="transparent")
        header_frame.pack(fill="x", pady=4)

        ctk.CTkLabel(
            header_frame, text="Headers:",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"]
        ).pack(side="left", padx=(0, 8))

        ctk.CTkOptionMenu(
            header_frame, variable=self.header_normalize,
            values=["none", "trim", "lowercase", "snake_case"],
            font=ctk.CTkFont(size=11), height=28, width=120,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            dropdown_fg_color=COLORS["bg_secondary"]
        ).pack(side="left")

        # Empty value replacement
        empty_frame = ctk.CTkFrame(opts_frame, fg_color="transparent")
        empty_frame.pack(fill="x", pady=4)

        ctk.CTkLabel(
            empty_frame, text="Replace empty cells with:",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"]
        ).pack(side="left", padx=(0, 8))

        ctk.CTkEntry(
            empty_frame, textvariable=self.empty_value,
            placeholder_text="(leave blank)",
            font=ctk.CTkFont(size=11), height=28, width=120,
            fg_color=COLORS["bg_dark"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"]
        ).pack(side="left")

        # Per-column transforms section
        sep = ctk.CTkFrame(self, fg_color=COLORS["border"], height=1)
        sep.pack(fill="x", padx=12, pady=(8, 4))

        col_header = ctk.CTkFrame(self, fg_color="transparent")
        col_header.pack(fill="x", padx=12, pady=(4, 4))

        ctk.CTkLabel(
            col_header, text="Per-Column Transforms",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left")

        # Per-column transform list
        list_frame = ctk.CTkFrame(self, fg_color=COLORS["bg_dark"], corner_radius=6)
        list_frame.pack(fill="both", expand=True, padx=12, pady=(0, 4))

        self.transforms_frame = ctk.CTkScrollableFrame(
            list_frame, fg_color="transparent",
            scrollbar_button_color=COLORS["bg_tertiary"]
        )
        self.transforms_frame.pack(fill="both", expand=True, padx=4, pady=4)

        self._refresh_transforms()

        # Buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=12, pady=(0, 12))

        ctk.CTkButton(
            btn_frame, text="+ Add Transform", font=ctk.CTkFont(size=11),
            height=28, fg_color=COLORS["accent_purple"],
            hover_color=COLORS["accent_purple_hover"],
            corner_radius=4, command=self._add_transform
        ).pack(side="left")

        ctk.CTkButton(
            btn_frame, text="Clear All", font=ctk.CTkFont(size=11),
            height=28, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["accent_red"],
            text_color=COLORS["text_secondary"],
            corner_radius=4, command=self._clear_transforms
        ).pack(side="right")

    def set_columns(self, columns: list[str]):
        self.columns = columns

    def _add_transform(self):
        if not self.columns:
            return
        self.column_transforms.append((self.columns[0], "trim", "", ""))
        self._refresh_transforms()

    def _remove_transform(self, idx: int):
        if 0 <= idx < len(self.column_transforms):
            self.column_transforms.pop(idx)
            self._refresh_transforms()

    def _clear_transforms(self):
        self.column_transforms.clear()
        self._refresh_transforms()

    def _update_transform(self, idx: int, **kwargs):
        if 0 <= idx < len(self.column_transforms):
            current = list(self.column_transforms[idx])
            if "column" in kwargs:
                current[0] = kwargs["column"]
            if "transform_type" in kwargs:
                current[1] = kwargs["transform_type"]
            if "arg1" in kwargs:
                if len(current) > 2:
                    current[2] = kwargs["arg1"]
                else:
                    current.append(kwargs["arg1"])
            if "arg2" in kwargs:
                if len(current) > 3:
                    current[3] = kwargs["arg2"]
                else:
                    current.append(kwargs["arg2"])
            self.column_transforms[idx] = tuple(current)

    def _refresh_transforms(self):
        for w in self.transforms_frame.winfo_children():
            w.destroy()

        if not self.column_transforms:
            ctk.CTkLabel(
                self.transforms_frame,
                text="No per-column transforms defined",
                font=ctk.CTkFont(size=11),
                text_color=COLORS["text_muted"]
            ).pack(pady=10)
        else:
            for idx, ct in enumerate(self.column_transforms):
                self._create_transform_row(idx, ct)

    def _create_transform_row(self, idx: int, ct: tuple):
        frame = ctk.CTkFrame(self.transforms_frame, fg_color=COLORS["bg_secondary"],
                             corner_radius=4)
        frame.pack(fill="x", pady=2)

        row1 = ctk.CTkFrame(frame, fg_color="transparent")
        row1.pack(fill="x", padx=8, pady=(6, 2))

        col_var = StringVar(value=ct[0])
        ctk.CTkOptionMenu(
            row1, variable=col_var,
            values=self.columns if self.columns else ["(no columns)"],
            font=ctk.CTkFont(size=10), height=26, width=120,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            dropdown_fg_color=COLORS["bg_secondary"],
            command=lambda v, i=idx: self._update_transform(i, column=v)
        ).pack(side="left", padx=(0, 4))

        type_display = {t: n for t, n in self.COLUMN_TRANSFORM_TYPES}
        type_var = StringVar(value=type_display.get(ct[1], ct[1]))
        ctk.CTkOptionMenu(
            row1, variable=type_var,
            values=[n for _, n in self.COLUMN_TRANSFORM_TYPES],
            font=ctk.CTkFont(size=10), height=26, width=100,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            dropdown_fg_color=COLORS["bg_secondary"],
            command=lambda v, i=idx: self._update_transform(
                i, transform_type=next(
                    (t for t, n in self.COLUMN_TRANSFORM_TYPES if n == v), v
                ))
        ).pack(side="left", padx=(0, 4))

        ctk.CTkButton(
            row1, text="✕", font=ctk.CTkFont(size=10),
            width=24, height=24, fg_color="transparent",
            hover_color=COLORS["accent_red"],
            text_color=COLORS["text_muted"], corner_radius=4,
            command=lambda i=idx: self._remove_transform(i)
        ).pack(side="right")

        # Show arg fields for replace/regex_replace
        if ct[1] in ("replace", "regex_replace"):
            row2 = ctk.CTkFrame(frame, fg_color="transparent")
            row2.pack(fill="x", padx=8, pady=(2, 6))

            search_entry = ctk.CTkEntry(
                row2, placeholder_text="Search...",
                font=ctk.CTkFont(size=10), height=26, width=120,
                fg_color=COLORS["bg_dark"],
                border_color=COLORS["border"],
                text_color=COLORS["text_primary"]
            )
            search_entry.pack(side="left", padx=(0, 4))
            if len(ct) > 2:
                search_entry.insert(0, ct[2])
            search_entry.bind("<KeyRelease>",
                              lambda e, i=idx: self._update_transform(i, arg1=e.widget.get()))

            replace_entry = ctk.CTkEntry(
                row2, placeholder_text="Replace with...",
                font=ctk.CTkFont(size=10), height=26, width=120,
                fg_color=COLORS["bg_dark"],
                border_color=COLORS["border"],
                text_color=COLORS["text_primary"]
            )
            replace_entry.pack(side="left")
            if len(ct) > 3:
                replace_entry.insert(0, ct[3])
            replace_entry.bind("<KeyRelease>",
                               lambda e, i=idx: self._update_transform(i, arg2=e.widget.get()))

    def get_config(self) -> tuple[bool, str, str]:
        return self.trim_whitespace.get(), self.case_transform.get(), self.empty_value.get()

    def get_header_normalize(self) -> str:
        return self.header_normalize.get()

    def get_column_transforms(self) -> list:
        return [ct for ct in self.column_transforms]


class OutputPanel(ctk.CTkFrame):
    """Output configuration."""
    
    def __init__(self, master, **kwargs):
        if "fg_color" not in kwargs:
            kwargs["fg_color"] = COLORS["bg_secondary"]
        if "corner_radius" not in kwargs:
            kwargs["corner_radius"] = 8
        super().__init__(master, **kwargs)
        
        self.delimiter = StringVar(value=",")
        self.encoding = StringVar(value="utf-8")
        self.quoting = StringVar(value="minimal")
        self.include_header = BooleanVar(value=True)
        self.line_ending = StringVar(value="auto")
        self.output_path = StringVar(value="")
        
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=12, pady=(12, 8))
        
        ctk.CTkLabel(
            header, text="💾 Output Settings",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left")
        
        # Options grid
        opts_frame = ctk.CTkFrame(self, fg_color="transparent")
        opts_frame.pack(fill="x", padx=12, pady=(0, 8))
        
        # Row 1: Delimiter and Encoding
        row1 = ctk.CTkFrame(opts_frame, fg_color="transparent")
        row1.pack(fill="x", pady=4)
        
        ctk.CTkLabel(row1, text="Delimiter:", font=ctk.CTkFont(size=11),
                     text_color=COLORS["text_secondary"], width=70, anchor="w").pack(side="left")
        ctk.CTkOptionMenu(
            row1, variable=self.delimiter,
            values=[",", ";", "\\t (Tab)", "|"],
            font=ctk.CTkFont(size=11), height=28, width=100,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            dropdown_fg_color=COLORS["bg_secondary"]
        ).pack(side="left", padx=(0, 20))
        
        ctk.CTkLabel(row1, text="Encoding:", font=ctk.CTkFont(size=11),
                     text_color=COLORS["text_secondary"], width=70, anchor="w").pack(side="left")
        ctk.CTkOptionMenu(
            row1, variable=self.encoding,
            values=["utf-8", "utf-16", "latin-1", "cp1252"],
            font=ctk.CTkFont(size=11), height=28, width=100,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            dropdown_fg_color=COLORS["bg_secondary"]
        ).pack(side="left")
        
        # Row 2: Quoting and Line ending
        row2 = ctk.CTkFrame(opts_frame, fg_color="transparent")
        row2.pack(fill="x", pady=4)
        
        ctk.CTkLabel(row2, text="Quoting:", font=ctk.CTkFont(size=11),
                     text_color=COLORS["text_secondary"], width=70, anchor="w").pack(side="left")
        ctk.CTkOptionMenu(
            row2, variable=self.quoting,
            values=["minimal", "all", "nonnumeric", "none"],
            font=ctk.CTkFont(size=11), height=28, width=100,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            dropdown_fg_color=COLORS["bg_secondary"]
        ).pack(side="left", padx=(0, 20))
        
        ctk.CTkLabel(row2, text="Line End:", font=ctk.CTkFont(size=11),
                     text_color=COLORS["text_secondary"], width=70, anchor="w").pack(side="left")
        ctk.CTkOptionMenu(
            row2, variable=self.line_ending,
            values=["auto", "unix (LF)", "windows (CRLF)"],
            font=ctk.CTkFont(size=11), height=28, width=100,
            fg_color=COLORS["bg_dark"],
            button_color=COLORS["bg_tertiary"],
            dropdown_fg_color=COLORS["bg_secondary"]
        ).pack(side="left")
        
        # Include header checkbox
        ctk.CTkCheckBox(
            opts_frame, text="Include header row",
            variable=self.include_header,
            font=ctk.CTkFont(size=11),
            fg_color=COLORS["accent_blue"],
            text_color=COLORS["text_secondary"]
        ).pack(anchor="w", pady=(8, 4))
        
        # Output path
        path_frame = ctk.CTkFrame(self, fg_color="transparent")
        path_frame.pack(fill="x", padx=12, pady=(0, 12))
        
        ctk.CTkLabel(
            path_frame, text="Output File:",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_secondary"]
        ).pack(anchor="w", pady=(0, 4))
        
        path_row = ctk.CTkFrame(path_frame, fg_color="transparent")
        path_row.pack(fill="x")
        
        self.path_entry = ctk.CTkEntry(
            path_row, textvariable=self.output_path,
            placeholder_text="Select output file...",
            font=ctk.CTkFont(size=11), height=32,
            fg_color=COLORS["bg_dark"],
            border_color=COLORS["border"],
            text_color=COLORS["text_primary"]
        )
        self.path_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        
        ctk.CTkButton(
            path_row, text="Browse", font=ctk.CTkFont(size=11),
            height=32, width=70, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_hover"],
            text_color=COLORS["text_primary"],
            corner_radius=6, command=self._browse
        ).pack(side="left")
    
    def _browse(self):
        file = filedialog.asksaveasfilename(
            title="Save As",
            defaultextension=".csv",
            filetypes=[
                ("CSV Files", "*.csv"),
                ("TSV Files", "*.tsv"),
                ("Excel Files", "*.xlsx"),
                ("Parquet Files", "*.parquet"),
                ("Text Files", "*.txt"),
                ("All Files", "*.*"),
            ]
        )
        if file:
            self.output_path.set(file)
    
    def get_config(self) -> dict:
        delim = self.delimiter.get()
        if delim == "\\t (Tab)":
            delim = "\t"
        
        line_end = self.line_ending.get()
        if "unix" in line_end:
            line_end = "unix"
        elif "windows" in line_end:
            line_end = "windows"
        else:
            line_end = "auto"
        
        return {
            "delimiter": delim,
            "encoding": self.encoding.get(),
            "quoting": self.quoting.get(),
            "include_header": self.include_header.get(),
            "line_ending": line_end,
            "output_path": self.output_path.get()
        }


class LogPanel(ctk.CTkFrame):
    """Processing log display."""
    
    def __init__(self, master, **kwargs):
        if "fg_color" not in kwargs:
            kwargs["fg_color"] = COLORS["bg_secondary"]
        if "corner_radius" not in kwargs:
            kwargs["corner_radius"] = 8
        super().__init__(master, **kwargs)
        
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=12, pady=(12, 8))
        
        ctk.CTkLabel(
            header, text="📋 Processing Log",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(side="left")
        
        ctk.CTkButton(
            header, text="Clear", font=ctk.CTkFont(size=10),
            height=24, width=50, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=4, command=self.clear
        ).pack(side="right")
        
        # Log text
        self.log_text = ctk.CTkTextbox(
            self, fg_color=COLORS["bg_dark"],
            text_color=COLORS["text_primary"],
            font=ctk.CTkFont(family="Consolas", size=11),
            corner_radius=6
        )
        self.log_text.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self.log_text.configure(state="disabled")
    
    def log(self, message: str, level: str = "info"):
        self.log_text.configure(state="normal")
        timestamp = datetime.now().strftime("%H:%M:%S")
        prefix = {"info": "│", "success": "✓", "warning": "⚠", "error": "✗"}.get(level, "│")
        self.log_text.insert(END, f"[{timestamp}] {prefix} {message}\n")
        self.log_text.see(END)
        self.log_text.configure(state="disabled")
    
    def clear(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", END)
        self.log_text.configure(state="disabled")


class StatsPanel(ctk.CTkFrame):
    """Processing statistics display."""
    
    def __init__(self, master, **kwargs):
        if "fg_color" not in kwargs:
            kwargs["fg_color"] = COLORS["bg_secondary"]
        if "corner_radius" not in kwargs:
            kwargs["corner_radius"] = 8
        super().__init__(master, **kwargs)
        
        self.labels = {}
        
        stats_config = [
            ("files_processed", "Files Processed", COLORS["accent_blue"]),
            ("total_rows_read", "Rows Read", COLORS["text_primary"]),
            ("rows_filtered", "Rows Filtered", COLORS["accent_orange"]),
            ("duplicates_removed", "Duplicates Removed", COLORS["accent_purple"]),
            ("final_row_count", "Final Rows", COLORS["accent_green"]),
        ]
        
        for i, (key, label, color) in enumerate(stats_config):
            frame = ctk.CTkFrame(self, fg_color="transparent")
            frame.pack(fill="x", padx=12, pady=(12 if i == 0 else 2, 2 if i < 4 else 12))
            
            ctk.CTkLabel(
                frame, text=label, font=ctk.CTkFont(size=11),
                text_color=COLORS["text_muted"]
            ).pack(side="left")
            
            val_label = ctk.CTkLabel(
                frame, text="—", font=ctk.CTkFont(size=12, weight="bold"),
                text_color=color
            )
            val_label.pack(side="right")
            self.labels[key] = val_label
    
    def update(self, stats: ProcessingStats):
        for key, label in self.labels.items():
            val = getattr(stats, key, 0)
            label.configure(text=f"{val:,}")
    
    def reset(self):
        for label in self.labels.values():
            label.configure(text="—")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN APPLICATION
# ══════════════════════════════════════════════════════════════════════════════

class CSVPowerToolApp:
    """Main application."""
    
    def __init__(self):
        if DND_AVAILABLE:
            self.root = TkinterDnD.Tk()
        else:
            self.root = ctk.CTk()
        
        self.root.title(APP_NAME)
        self.root.geometry("1200x850")
        self.root.minsize(1000, 700)
        
        ctk.set_appearance_mode("dark")
        self.root.configure(bg=COLORS["bg_dark"])
        
        self.engine: CSVEngine = None
        self.processing = False
        
        self._build_ui()
        
        if DND_AVAILABLE:
            self._setup_dnd()
    
    def _build_ui(self):
        # Main container
        main = ctk.CTkFrame(self.root, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=16, pady=16)
        
        # Header
        header = ctk.CTkFrame(main, fg_color="transparent")
        header.pack(fill="x", pady=(0, 12))
        
        title_frame = ctk.CTkFrame(header, fg_color="transparent")
        title_frame.pack(side="left")
        
        ctk.CTkLabel(
            title_frame, text=APP_NAME,
            font=ctk.CTkFont(size=26, weight="bold"),
            text_color=COLORS["text_primary"]
        ).pack(anchor="w")
        
        ctk.CTkLabel(
            title_frame, text="Combine • Filter • Transform • Deduplicate • Export",
            font=ctk.CTkFont(size=12),
            text_color=COLORS["text_muted"]
        ).pack(anchor="w")
        
        # Version
        ver_badge = ctk.CTkFrame(header, fg_color=COLORS["bg_tertiary"], corner_radius=12)
        ver_badge.pack(side="right")
        ctk.CTkLabel(ver_badge, text=f"v{APP_VERSION}", font=ctk.CTkFont(size=11),
                     text_color=COLORS["text_muted"]).pack(padx=12, pady=4)
        
        # Content: 3 columns
        content = ctk.CTkFrame(main, fg_color="transparent")
        content.pack(fill="both", expand=True)
        
        # Left column: Files
        left_col = ctk.CTkFrame(content, fg_color="transparent", width=280)
        left_col.pack(side="left", fill="both", padx=(0, 8))
        left_col.pack_propagate(False)
        
        self.file_panel = FileListPanel(left_col, on_change=self._on_files_changed)
        self.file_panel.pack(fill="both", expand=True)
        
        # Middle column: Configuration tabs
        mid_col = ctk.CTkFrame(content, fg_color="transparent")
        mid_col.pack(side="left", fill="both", expand=True, padx=8)
        
        self.tabview = ctk.CTkTabview(
            mid_col, fg_color=COLORS["bg_secondary"],
            segmented_button_fg_color=COLORS["bg_tertiary"],
            segmented_button_selected_color=COLORS["accent_blue"],
            segmented_button_selected_hover_color=COLORS["accent_blue_hover"],
            segmented_button_unselected_color=COLORS["bg_tertiary"],
            segmented_button_unselected_hover_color=COLORS["bg_hover"],
            text_color=COLORS["text_primary"],
            corner_radius=8
        )
        self.tabview.pack(fill="both", expand=True)
        
        # Create tabs
        tab_columns = self.tabview.add("Columns")
        tab_sort = self.tabview.add("Sort")
        tab_dedupe = self.tabview.add("Dedupe")
        tab_filter = self.tabview.add("Filter")
        tab_transform = self.tabview.add("Transform")
        tab_output = self.tabview.add("Output")
        
        # Tab contents
        self.column_panel = ColumnPanel(tab_columns, fg_color="transparent")
        self.column_panel.pack(fill="both", expand=True)
        
        self.sort_panel = SortPanel(tab_sort, fg_color="transparent")
        self.sort_panel.pack(fill="both", expand=True)
        
        self.dedupe_panel = DedupePanel(tab_dedupe, fg_color="transparent")
        self.dedupe_panel.pack(fill="both", expand=True)
        
        self.filter_panel = FilterPanel(tab_filter, fg_color="transparent")
        self.filter_panel.pack(fill="both", expand=True)
        
        self.transform_panel = TransformPanel(tab_transform, fg_color="transparent")
        self.transform_panel.pack(fill="both", expand=True)
        
        self.output_panel = OutputPanel(tab_output, fg_color="transparent")
        self.output_panel.pack(fill="both", expand=True)
        
        # Right column: Log and Stats
        right_col = ctk.CTkFrame(content, fg_color="transparent", width=300)
        right_col.pack(side="right", fill="both", padx=(8, 0))
        right_col.pack_propagate(False)
        
        self.stats_panel = StatsPanel(right_col)
        self.stats_panel.pack(fill="x", pady=(0, 8))
        
        self.log_panel = LogPanel(right_col)
        self.log_panel.pack(fill="both", expand=True)
        
        # Bottom: Progress and buttons
        bottom = ctk.CTkFrame(main, fg_color="transparent")
        bottom.pack(fill="x", pady=(12, 0))
        
        # Progress
        progress_frame = ctk.CTkFrame(bottom, fg_color="transparent")
        progress_frame.pack(fill="x", pady=(0, 8))
        
        self.progress_label = ctk.CTkLabel(
            progress_frame, text="Ready",
            font=ctk.CTkFont(size=11),
            text_color=COLORS["text_muted"]
        )
        self.progress_label.pack(anchor="w", pady=(0, 4))
        
        self.progress_bar = ctk.CTkProgressBar(
            progress_frame, height=8,
            fg_color=COLORS["bg_tertiary"],
            progress_color=COLORS["accent_green"],
            corner_radius=4
        )
        self.progress_bar.pack(fill="x")
        self.progress_bar.set(0)
        
        # Buttons
        btn_frame = ctk.CTkFrame(bottom, fg_color="transparent")
        btn_frame.pack(fill="x")
        
        # Preset buttons
        preset_frame = ctk.CTkFrame(btn_frame, fg_color="transparent")
        preset_frame.pack(side="left")
        
        ctk.CTkButton(
            preset_frame, text="💾 Save Config", font=ctk.CTkFont(size=11),
            height=36, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=6, command=self._save_config
        ).pack(side="left", padx=(0, 4))
        
        ctk.CTkButton(
            preset_frame, text="📂 Load Config", font=ctk.CTkFont(size=11),
            height=36, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_hover"],
            text_color=COLORS["text_secondary"],
            corner_radius=6, command=self._load_config
        ).pack(side="left")
        
        # Action buttons
        action_frame = ctk.CTkFrame(btn_frame, fg_color="transparent")
        action_frame.pack(side="right")
        
        self.cancel_btn = ctk.CTkButton(
            action_frame, text="Cancel", font=ctk.CTkFont(size=13),
            height=42, width=100, fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["accent_red"],
            text_color=COLORS["text_primary"],
            corner_radius=8, command=self._cancel,
            state="disabled"
        )
        self.cancel_btn.pack(side="left", padx=(0, 8))
        
        self.process_btn = ctk.CTkButton(
            action_frame, text="▶  Process Files",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=42, width=160,
            fg_color=COLORS["accent_green"],
            hover_color=COLORS["accent_green_hover"],
            corner_radius=8, command=self._process
        )
        self.process_btn.pack(side="left")
    
    def _setup_dnd(self):
        def drop(event):
            files = self.root.tk.splitlist(event.data)
            added = self.file_panel.add_files([Path(f) for f in files])
            if added:
                self.log_panel.log(f"Added {added} file(s) via drag & drop", "info")
        
        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind('<<Drop>>', drop)
    
    def _on_files_changed(self):
        if self.file_panel.files:
            columns = CSVEngine(ProcessingConfig()).discover_columns(self.file_panel.files)
            self.column_panel.set_columns(columns)
            self.sort_panel.set_columns(columns)
            self.dedupe_panel.set_columns(columns)
            self.filter_panel.set_columns(columns)
            self.transform_panel.set_columns(columns)
    
    def _update_progress(self, value: float, status: str):
        self.progress_bar.set(value / 100)
        self.progress_label.configure(text=status)
        self.root.update_idletasks()
    
    def _build_config(self) -> ProcessingConfig:
        config = ProcessingConfig()
        
        # Columns
        mode, selected = self.column_panel.get_config()
        config.columns_mode = mode
        config.selected_columns = selected
        
        # Sort
        enabled, rules, case_sens, numeric = self.sort_panel.get_config()
        config.sort_enabled = enabled
        config.sort_columns = rules
        config.sort_case_sensitive = case_sens
        config.sort_numeric_aware = numeric
        
        # Dedupe
        (
            enabled,
            columns,
            keep,
            fuzzy_enabled,
            fuzzy_threshold,
            aggregate_mode,
            aggregate_separator,
        ) = self.dedupe_panel.get_config()
        config.dedupe_enabled = enabled
        config.dedupe_columns = columns
        config.dedupe_keep = keep
        config.dedupe_fuzzy_enabled = fuzzy_enabled
        config.dedupe_fuzzy_threshold = fuzzy_threshold
        config.dedupe_aggregate_mode = aggregate_mode
        config.dedupe_aggregate_separator = aggregate_separator
        
        # Filter
        filters, logic = self.filter_panel.get_config()
        config.filters = filters
        config.filter_logic = logic
        
        # Transform
        trim, case, empty = self.transform_panel.get_config()
        config.trim_whitespace = trim
        config.case_transform = case
        config.empty_value = empty
        config.header_normalize = self.transform_panel.get_header_normalize()
        config.column_transforms = self.transform_panel.get_column_transforms()

        # Output
        output_config = self.output_panel.get_config()
        config.output_delimiter = output_config["delimiter"]
        config.output_encoding = output_config["encoding"]
        config.output_quoting = output_config["quoting"]
        config.include_header = output_config["include_header"]
        config.line_ending = output_config["line_ending"]

        return config
    
    def _process(self):
        if not self.file_panel.files:
            self.log_panel.log("No files selected", "error")
            return
        
        output_config = self.output_panel.get_config()
        output_path = output_config["output_path"]
        
        if not output_path:
            output_path = str(Path.home() / "combined_output.csv")
            self.output_panel.output_path.set(output_path)
        
        self.processing = True
        self._set_ui_state(True)
        self.stats_panel.reset()
        self.progress_bar.set(0)
        self.log_panel.clear()
        self.log_panel.log("Starting processing...", "info")
        
        config = self._build_config()
        
        def run():
            self.engine = CSVEngine(
                config,
                progress_callback=lambda v, s: self.root.after(0, lambda: self._update_progress(v, s)),
                log_callback=lambda m, l: self.root.after(0, lambda: self.log_panel.log(m, l))
            )
            
            stats = self.engine.process(self.file_panel.files, Path(output_path))
            self.root.after(0, lambda: self._complete(stats))
        
        threading.Thread(target=run, daemon=True).start()
    
    def _cancel(self):
        if self.engine:
            self.engine.cancel()
            self.log_panel.log("Cancelling...", "warning")
    
    def _complete(self, stats: ProcessingStats):
        self.processing = False
        self._set_ui_state(False)
        self.stats_panel.update(stats)
        
        if stats.final_row_count > 0:
            self.log_panel.log(f"✓ Complete! {stats.final_row_count:,} rows saved", "success")
        else:
            self.log_panel.log("Processing completed with no output", "warning")
    
    def _set_ui_state(self, processing: bool):
        state = "disabled" if processing else "normal"
        self.process_btn.configure(state=state)
        self.cancel_btn.configure(state="normal" if processing else "disabled")
    
    def _save_config(self):
        file = filedialog.asksaveasfilename(
            title="Save Configuration",
            defaultextension=".json",
            filetypes=[("JSON Files", "*.json")]
        )
        if file:
            config = self._build_config()
            data = {
                "columns_mode": config.columns_mode,
                "selected_columns": config.selected_columns,
                "sort_enabled": config.sort_enabled,
                "sort_columns": config.sort_columns,
                "sort_case_sensitive": config.sort_case_sensitive,
                "sort_numeric_aware": config.sort_numeric_aware,
                "dedupe_enabled": config.dedupe_enabled,
                "dedupe_columns": config.dedupe_columns,
                "dedupe_keep": config.dedupe_keep,
                "dedupe_fuzzy_enabled": config.dedupe_fuzzy_enabled,
                "dedupe_fuzzy_threshold": config.dedupe_fuzzy_threshold,
                "dedupe_aggregate_mode": config.dedupe_aggregate_mode,
                "dedupe_aggregate_separator": config.dedupe_aggregate_separator,
                "filters": config.filters,
                "filter_logic": config.filter_logic,
                "trim_whitespace": config.trim_whitespace,
                "case_transform": config.case_transform,
                "empty_value": config.empty_value,
                "header_normalize": config.header_normalize,
                "column_transforms": config.column_transforms,
                "schema_mode": config.schema_mode,
                "output_delimiter": config.output_delimiter,
                "output_encoding": config.output_encoding,
                "output_quoting": config.output_quoting,
                "include_header": config.include_header,
                "line_ending": config.line_ending,
            }
            with open(file, 'w') as f:
                json.dump(data, f, indent=2)
            self.log_panel.log(f"Configuration saved: {Path(file).name}", "success")
    
    def _load_config(self):
        file = filedialog.askopenfilename(
            title="Load Configuration",
            filetypes=[("JSON Files", "*.json")]
        )
        if file:
            try:
                with open(file, 'r') as f:
                    data = json.load(f)
                
                # Apply to panels
                self.column_panel.mode.set(data.get("columns_mode", "all"))
                self.column_panel.selected = set(data.get("selected_columns", []))
                
                self.sort_panel.enabled.set(data.get("sort_enabled", False))
                self.sort_panel.sort_rules = data.get("sort_columns", [])
                self.sort_panel.case_sensitive.set(data.get("sort_case_sensitive", False))
                self.sort_panel.numeric_aware.set(data.get("sort_numeric_aware", True))
                self.sort_panel._refresh_rules()
                
                self.dedupe_panel.enabled.set(data.get("dedupe_enabled", True))
                self.dedupe_panel.selected_columns = set(data.get("dedupe_columns", []))
                self.dedupe_panel.keep_mode.set(data.get("dedupe_keep", "first"))
                self.dedupe_panel.fuzzy_enabled.set(data.get("dedupe_fuzzy_enabled", False))
                self.dedupe_panel.fuzzy_threshold.set(data.get("dedupe_fuzzy_threshold", 90))
                self.dedupe_panel.aggregate_mode.set(data.get("dedupe_aggregate_mode", "none"))
                self.dedupe_panel.aggregate_separator.set(data.get("dedupe_aggregate_separator", "; "))
                
                self.filter_panel.filters = data.get("filters", [])
                self.filter_panel.logic.set(data.get("filter_logic", "and"))
                self.filter_panel._refresh()
                
                self.transform_panel.trim_whitespace.set(data.get("trim_whitespace", True))
                self.transform_panel.case_transform.set(data.get("case_transform", "none"))
                self.transform_panel.empty_value.set(data.get("empty_value", ""))
                self.transform_panel.header_normalize.set(data.get("header_normalize", "none"))
                self.transform_panel.column_transforms = data.get("column_transforms", [])
                self.transform_panel._refresh_transforms()

                delim = data.get("output_delimiter", ",")
                if delim == "\t":
                    delim = "\\t (Tab)"
                self.output_panel.delimiter.set(delim)
                self.output_panel.encoding.set(data.get("output_encoding", "utf-8"))
                self.output_panel.quoting.set(data.get("output_quoting", "minimal"))
                self.output_panel.include_header.set(data.get("include_header", True))
                
                self.log_panel.log(f"Configuration loaded: {Path(file).name}", "success")
                
            except Exception as e:
                self.log_panel.log(f"Error loading config: {e}", "error")
    
    def run(self):
        self.root.mainloop()


# ══════════════════════════════════════════════════════════════════════════════
# CLI MODE
# ══════════════════════════════════════════════════════════════════════════════

def cli_main():
    """Headless CLI mode: csv_power_tool --config preset.json --inputs *.csv --output combined.csv"""
    import argparse
    import glob as globmod

    parser = argparse.ArgumentParser(
        prog="csv_power_tool",
        description="CSV Power Tool - Professional-grade CSV combiner and processor (CLI mode)",
    )
    parser.add_argument("--config", "-c", help="JSON preset configuration file")
    parser.add_argument("--inputs", "-i", nargs="+", required=True,
                        help="Input files, folders, or glob patterns (e.g. *.csv)")
    parser.add_argument("--output", "-o", required=True, help="Output file path")
    parser.add_argument("--delimiter", "-d", default=None, help="Output delimiter")
    parser.add_argument("--encoding", "-e", default=None, help="Output encoding")
    parser.add_argument("--no-header", action="store_true", help="Exclude header row")
    parser.add_argument("--no-dedupe", action="store_true", help="Disable deduplication")
    parser.add_argument("--fuzzy-dedupe-threshold", type=int, metavar="50-100",
                        help="Enable fuzzy dedupe at the supplied similarity threshold")
    parser.add_argument("--dedupe-aggregate", choices=["max", "min", "sum", "concat"],
                        help="Aggregate non-key duplicate columns with the selected mode")
    parser.add_argument("--dedupe-aggregate-separator", default="; ",
                        help="Separator for --dedupe-aggregate concat")
    parser.add_argument("--filter", action="append", metavar="COL:OP:VALUE",
                        help="Add a filter rule. Operators include between, fuzzy, regex, contains")
    parser.add_argument("--sort", nargs="+", metavar="COL[:asc|desc]",
                        help="Sort by columns, e.g. name:asc age:desc")
    parser.add_argument("--columns", nargs="+", help="Include only these columns")
    parser.add_argument("--exclude-columns", nargs="+", help="Exclude these columns")
    parser.add_argument("--header-normalize", choices=["none", "trim", "lowercase", "snake_case"],
                        default="none", help="Header normalization mode")
    parser.add_argument("--schema-mode", choices=["union", "intersection", "first_file"],
                        default="union", help="Schema unification mode")
    parser.add_argument("--no-stream", action="store_true", help="Disable bounded-memory streaming path")
    parser.add_argument("--watch", action="store_true",
                        help="Watch inputs and re-run whenever matching files change")
    parser.add_argument("--watch-interval", type=float, default=2.0,
                        help="Polling interval for --watch, in seconds")
    parser.add_argument("--quiet", "-q", action="store_true", help="Suppress log output")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")
    parser.add_argument("--version", action="version", version=f"{APP_NAME} v{APP_VERSION}")

    args = parser.parse_args()

    def expand_inputs(patterns):
        input_files = []
        seen = set()
        for pattern in patterns:
            p = Path(pattern)
            candidates = []
            if p.is_dir():
                candidates = [child for child in p.rglob("*") if child.is_file()]
            else:
                expanded = globmod.glob(pattern, recursive=True)
                if expanded:
                    candidates = [Path(f) for f in expanded]
                elif p.exists():
                    candidates = [p]
                else:
                    print(f"Warning: No files matched: {pattern}", file=sys.stderr)

            for candidate in candidates:
                if candidate.suffix.lower() not in SUPPORTED_INPUT_SUFFIXES:
                    continue
                resolved = candidate.resolve()
                if resolved not in seen:
                    seen.add(resolved)
                    input_files.append(candidate)
        return sorted(input_files, key=lambda path: str(path).lower())

    # Build config
    config = ProcessingConfig()

    # Load preset if provided
    if args.config:
        try:
            with open(args.config, 'r') as f:
                data = json.load(f)
            for key, value in data.items():
                if hasattr(config, key):
                    setattr(config, key, value)
        except Exception as e:
            print(f"Error loading config: {e}", file=sys.stderr)
            sys.exit(3)

    # Apply CLI overrides
    if args.delimiter:
        delim = args.delimiter
        if delim.lower() == "tab" or delim == "\\t":
            delim = "\t"
        config.output_delimiter = delim

    if args.encoding:
        config.output_encoding = args.encoding

    if args.no_header:
        config.include_header = False

    if args.no_dedupe:
        config.dedupe_enabled = False

    if args.fuzzy_dedupe_threshold is not None:
        if not 50 <= args.fuzzy_dedupe_threshold <= 100:
            print("Error: --fuzzy-dedupe-threshold must be between 50 and 100", file=sys.stderr)
            sys.exit(3)
        config.dedupe_enabled = True
        config.dedupe_fuzzy_enabled = True
        config.dedupe_fuzzy_threshold = args.fuzzy_dedupe_threshold

    if args.dedupe_aggregate:
        config.dedupe_enabled = True
        config.dedupe_aggregate_mode = args.dedupe_aggregate
        config.dedupe_aggregate_separator = args.dedupe_aggregate_separator

    if args.columns:
        config.columns_mode = "select"
        config.selected_columns = args.columns

    if args.exclude_columns:
        config.columns_mode = "exclude"
        config.selected_columns = args.exclude_columns

    config.header_normalize = args.header_normalize
    config.schema_mode = args.schema_mode
    config.streaming_enabled = not args.no_stream

    if args.filter:
        for spec in args.filter:
            parts = spec.split(":", 2)
            if len(parts) != 3:
                print(f"Error: Invalid filter format: {spec}", file=sys.stderr)
                sys.exit(3)
            col, operator, value = parts
            if operator not in CSVEngine.FILTER_OPERATORS:
                print(f"Error: Unknown filter operator: {operator}", file=sys.stderr)
                sys.exit(3)
            config.filters.append((col, operator, value))

    if args.sort:
        config.sort_enabled = True
        config.sort_columns = []
        for s in args.sort:
            if ":" in s:
                col, direction = s.rsplit(":", 1)
                ascending = direction.lower() != "desc"
            else:
                col = s
                ascending = True
            config.sort_columns.append((col, ascending))

    # Log callback
    def log_msg(message, level="info"):
        if args.quiet:
            return
        prefix = {"info": "  ", "success": "OK", "warning": "!!", "error": "XX"}.get(level, "  ")
        print(f"[{prefix}] {message}", file=sys.stderr)

    def progress_msg(value, status):
        if args.verbose and not args.quiet:
            print(f"  ... {status} ({value:.0f}%)", file=sys.stderr)

    output_path = Path(args.output)

    def run_once(input_files):
        if not input_files:
            print("Error: No input files found", file=sys.stderr)
            return 1

        engine = CSVEngine(config, progress_callback=progress_msg, log_callback=log_msg)

        if not args.quiet:
            print(f"CSV Power Tool - Processing {len(input_files)} file(s)...", file=sys.stderr)

        stats = engine.process(input_files, output_path)

        if not args.quiet:
            print(f"\nResults:", file=sys.stderr)
            print(f"  Files processed:    {stats.files_processed}", file=sys.stderr)
            print(f"  Files skipped:      {stats.files_skipped}", file=sys.stderr)
            print(f"  Total rows read:    {stats.total_rows_read:,}", file=sys.stderr)
            print(f"  Rows filtered:      {stats.rows_filtered:,}", file=sys.stderr)
            print(f"  Duplicates removed: {stats.duplicates_removed:,}", file=sys.stderr)
            print(f"  Final row count:    {stats.final_row_count:,}", file=sys.stderr)
            print(f"  Output: {output_path}", file=sys.stderr)

        if stats.files_processed == 0:
            return 1
        if stats.errors:
            return 3
        return 0

    def input_signature(input_files):
        signature = []
        for path in input_files:
            try:
                stat = path.stat()
            except OSError:
                continue
            signature.append((str(path.resolve()), stat.st_mtime_ns, stat.st_size))
        return tuple(signature)

    if args.watch:
        last_signature = None
        last_exit = 0
        try:
            while True:
                watched_files = expand_inputs(args.inputs)
                signature = input_signature(watched_files)
                if signature != last_signature:
                    if watched_files:
                        last_exit = run_once(watched_files)
                    elif not args.quiet:
                        print("CSV Power Tool - waiting for input files...", file=sys.stderr)
                    last_signature = signature
                time.sleep(max(args.watch_interval, 0.1))
        except KeyboardInterrupt:
            sys.exit(last_exit)

    input_files = expand_inputs(args.inputs)
    sys.exit(run_once(input_files))


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # If CLI arguments are present, run headless; otherwise launch GUI
    if len(sys.argv) > 1:
        cli_main()
    else:
        if GUI_IMPORT_ERROR is not None:
            print(
                "GUI mode requires customtkinter. Install dependencies with: "
                "python -m pip install -r requirements.txt",
                file=sys.stderr,
            )
            sys.exit(4)
        app = CSVPowerToolApp()
        app.run()
